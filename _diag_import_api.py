from openpyxl import load_workbook
from pathlib import Path
p = Path('出口易物流推广报价表2026.04.22-1.xlsx')
wb = load_workbook(p, data_only=True, read_only=False)
ws = wb['美国商派方案']
print('max_row=', ws.max_row)
print('max_col=', ws.max_column)
print('merged_ranges=', len(ws.merged_cells.ranges))
wb.close()

from openpyxl import load_workbook
from pathlib import Path

wb = load_workbook(Path("出口易物流推广报价表2026.04.22-1.xlsx"), data_only=True, read_only=True)
ws = wb["出口易拉美专线"]
for r in range(1, 40):
    vals = []
    for c in range(1, 12):
        v = ws.cell(row=r, column=c).value
        s = "" if v is None else str(v).strip().replace("\n", "|")
        vals.append(s)
    row_text = "\t".join(vals)
    if any(x in row_text for x in ["国家", "重量", "运费", "处理费", "出口易", "SAX", "SAE"]):
        print(f"R{r}: {row_text}")
wb.close()

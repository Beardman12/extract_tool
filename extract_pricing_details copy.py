#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按产品线提取物流报价明细（支持合并单元格）

规则：
1. 识别标准表头：国家、重量段/KG、运费（RMB/KG）、处理费(RMB/票)、上网参考时效、尺寸限制
2. 表头上一行是更新时间，上上一行是产品代码行
3. 数据结束判定：在表头对应列中，若一行有 >=3 个空值（且已考虑合并单元格映射）并且不满足数据行特征，则判定当前产品线数据结束
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


EXPECTED_HEADERS = [
    "国家",
    "重量段/KG",
    "运费（RMB/KG）",
    "处理费(RMB/票)",
    "上网参考时效",
    "尺寸限制",
]

# 对表头做轻微归一化后匹配，兼容中英文括号/空格差异
HEADER_ALIASES = {
    "国家": "国家",
    "重量段kg": "重量段/KG",
    "重量段/kg": "重量段/KG",
    "重量段": "重量段/KG",
    "运费rmb/kg": "运费（RMB/KG）",
    "运费（rmb/kg）": "运费（RMB/KG）",
    "运费(rmb/kg)": "运费（RMB/KG）",
    "运费": "运费（RMB/KG）",
    "处理费rmb/票": "处理费(RMB/票)",
    "处理费（rmb/票）": "处理费(RMB/票)",
    "处理费(rmb/票)": "处理费(RMB/票)",
    "处理费": "处理费(RMB/票)",
    "上网参考时效": "上网参考时效",
    "尺寸限制": "尺寸限制",
}


@dataclass
class ProductLineTable:
    sheet_name: str
    product_line: str
    product_codes: List[str]
    update_time: str
    header_row: int
    data_start_row: int
    data_end_row: int
    headers: List[str]
    rows: List[Dict[str, object]]


class PricingExtractor:
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path, data_only=True)

    @staticmethod
    def _norm_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        text = text.replace("（", "(").replace("）", ")")
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @staticmethod
    def _display_text(value: object) -> str:
        return "" if value is None else str(value).strip()

    def _build_merged_lookup(self, ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
        lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    lookup[(row, col)] = (min_row, min_col)
        return lookup

    def _get_value(self, ws: Worksheet, row: int, col: int, merged_lookup: Dict[Tuple[int, int], Tuple[int, int]]) -> object:
        value = ws.cell(row=row, column=col).value
        if value is not None:
            return value

        anchor = merged_lookup.get((row, col))
        if anchor:
            anchor_row, anchor_col = anchor
            return ws.cell(row=anchor_row, column=anchor_col).value
        return None

    def _match_header_row(self, ws: Worksheet, row: int, merged_lookup: Dict[Tuple[int, int], Tuple[int, int]]) -> Optional[Tuple[int, List[str]]]:
        max_col = ws.max_column
        # 6列表头，扫描起始列
        for start_col in range(1, max_col - 5 + 1):
            cells = [self._get_value(ws, row, start_col + i, merged_lookup) for i in range(6)]
            normalized = [self._norm_text(v) for v in cells]
            mapped: List[str] = []
            ok = True
            for item in normalized:
                if item not in HEADER_ALIASES:
                    ok = False
                    break
                mapped.append(HEADER_ALIASES[item])

            if ok and mapped == EXPECTED_HEADERS:
                return start_col, mapped
        return None

    @staticmethod
    def _extract_product_codes(text: str) -> List[str]:
        if not text:
            return []

        codes = set()
        for code in re.findall(r"[\(（]([A-Z]{2,8})[\)）]", text):
            codes.add(code)

        for code in re.findall(r"\b([A-Z]{2,8})\b", text):
            codes.add(code)

        return sorted(codes)

    @staticmethod
    def _looks_like_update_time(text: str) -> bool:
        if not text:
            return False
        if "更新" in text and ("年" in text or ":" in text or "：" in text):
            return True
        if re.search(r"20\d{2}[./-]?\d{1,2}[./-]?\d{1,2}", text):
            return True
        return False

    @staticmethod
    def _looks_like_weight(text: str) -> bool:
        if not text:
            return False
        return bool(re.search(r"[<>≤≥Ww]|kg|KG", text))

    @staticmethod
    def _looks_like_price(value: object) -> bool:
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        text = str(value).strip()
        return bool(re.fullmatch(r"\d+(\.\d+)?", text))

    @staticmethod
    def _is_blank(value: object) -> bool:
        if value is None:
            return True
        return str(value).strip() == ""

    @staticmethod
    def _looks_like_next_product_line(text: str) -> bool:
        if not text:
            return False
        # 常见产品线名称内会带代码括号，如 (CTD)
        if re.search(r"[\(（][A-Z]{2,8}[\)）]", text):
            return True
        # 排除普通数据行
        if "出口易" in text and ("USPS" in text or "Gofo" in text or "SPX" in text):
            return True
        return False

    def _extract_one_table(
        self,
        ws: Worksheet,
        header_row: int,
        start_col: int,
        headers: List[str],
        merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
    ) -> ProductLineTable:
        product_line_text = self._display_text(self._get_value(ws, header_row - 2, start_col, merged_lookup)) if header_row >= 3 else ""
        update_time_text = self._display_text(self._get_value(ws, header_row - 1, start_col, merged_lookup)) if header_row >= 2 else ""

        # 若标准位置为空，向该行其它列兜底找第一个非空文本
        if not product_line_text and header_row >= 3:
            for c in range(1, ws.max_column + 1):
                val = self._display_text(self._get_value(ws, header_row - 2, c, merged_lookup))
                if val:
                    product_line_text = val
                    break

        if not update_time_text and header_row >= 2:
            for c in range(1, ws.max_column + 1):
                val = self._display_text(self._get_value(ws, header_row - 1, c, merged_lookup))
                if val and self._looks_like_update_time(val):
                    update_time_text = val
                    break

        codes = self._extract_product_codes(product_line_text)

        rows: List[Dict[str, object]] = []
        data_start = header_row + 1
        data_end = header_row

        row = data_start
        while row <= ws.max_row:
            # 新表头出现，当前表结束
            maybe_next_header = self._match_header_row(ws, row, merged_lookup)
            if maybe_next_header is not None:
                break

            values = [self._get_value(ws, row, start_col + i, merged_lookup) for i in range(6)]
            text_values = [self._display_text(v) for v in values]

            if all(self._is_blank(v) for v in values):
                # 全空直接结束
                break

            empty_count = sum(1 for v in values if self._is_blank(v))
            country, weight, freight, handling, eta, size_limit = values

            has_country = not self._is_blank(country)
            has_weight = self._looks_like_weight(self._display_text(weight))
            has_price = self._looks_like_price(freight) and self._looks_like_price(handling)
            row_text = " ".join(t for t in text_values if t)

            # 命中“下一产品线”文本，结束当前数据
            if self._looks_like_next_product_line(row_text) and not has_weight:
                break

            # 结束规则：>=3 空，并且该行不具备数据行特征
            if empty_count >= 3 and not (has_weight or has_price or has_country):
                break

            # 只保留真正的数据行，防止把服务说明等混入
            if has_weight or (has_country and has_price):
                row_item = {
                    "国家": self._display_text(country),
                    "重量段/KG": self._display_text(weight),
                    "运费（RMB/KG）": freight,
                    "处理费(RMB/票)": handling,
                    "上网参考时效": self._display_text(eta),
                    "尺寸限制": self._display_text(size_limit),
                    "行号": row,
                }
                rows.append(row_item)
                data_end = row
            else:
                # 非数据行时，一般视作当前产品线数据已经结束
                break

            row += 1

        if not rows:
            data_start = header_row + 1
            data_end = header_row

        return ProductLineTable(
            sheet_name=ws.title,
            product_line=product_line_text,
            product_codes=codes,
            update_time=update_time_text,
            header_row=header_row,
            data_start_row=data_start,
            data_end_row=data_end,
            headers=headers,
            rows=rows,
        )

    def extract(self) -> List[ProductLineTable]:
        all_tables: List[ProductLineTable] = []
        for ws in self.workbook.worksheets:
            merged_lookup = self._build_merged_lookup(ws)
            row = 1
            while row <= ws.max_row:
                header_info = self._match_header_row(ws, row, merged_lookup)
                if header_info is None:
                    row += 1
                    continue

                start_col, headers = header_info
                table = self._extract_one_table(ws, row, start_col, headers, merged_lookup)

                if table.rows:
                    all_tables.append(table)
                    row = max(table.data_end_row + 1, row + 1)
                else:
                    row += 1

        return all_tables


def tables_to_records(tables: List[ProductLineTable]) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for table in tables:
        for row in table.rows:
            record = {
                "sheet": table.sheet_name,
                "产品线": table.product_line,
                "产品代码": ",".join(table.product_codes),
                "更新时间": table.update_time,
                "header_row": table.header_row,
            }
            record.update(row)
            records.append(record)
    return records


def write_outputs(tables: List[ProductLineTable], output_json: Path, output_xlsx: Path) -> None:
    output_json.parent.mkdir(parents=True, exist_ok=True)

    payload = []
    for table in tables:
        payload.append(
            {
                "sheet": table.sheet_name,
                "产品线": table.product_line,
                "产品代码": table.product_codes,
                "更新时间": table.update_time,
                "header_row": table.header_row,
                "data_start_row": table.data_start_row,
                "data_end_row": table.data_end_row,
                "headers": table.headers,
                "rows": table.rows,
            }
        )

    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # 延迟导入 pandas，避免作为核心解析依赖
    import pandas as pd  # pylint: disable=import-outside-toplevel

    df = pd.DataFrame(tables_to_records(tables))
    df.to_excel(output_xlsx, index=False)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="提取报价表中的每个产品线价格明细")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("出口易物流推广报价表2026.04.22.xlsx"),
        help="输入 Excel 文件路径",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("output") / "pricing_details.json",
        help="JSON 输出路径",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("output") / "pricing_details.xlsx",
        help="Excel 输出路径",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()

    extractor = PricingExtractor(args.input)
    tables = extractor.extract()

    write_outputs(tables, args.output_json, args.output_xlsx)

    print(f"识别到产品线数量: {len(tables)}")
    for idx, table in enumerate(tables, start=1):
        codes = ",".join(table.product_codes) if table.product_codes else "(无代码)"
        print(
            f"[{idx}] sheet={table.sheet_name} | 代码={codes} | 更新时间={table.update_time} | 数据行={len(table.rows)}"
        )

    print(f"JSON输出: {args.output_json}")
    print(f"Excel输出: {args.output_xlsx}")


if __name__ == "__main__":
    main()

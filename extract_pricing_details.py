#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按产品线提取物流报价明细（支持合并单元格）

规则：
1. 识别标准表头：国家、重量段/KG、运费（RMB/KG）、处理费(RMB/票)、上网参考时效、尺寸限制
2. 表头上一行是更新时间，上上一行是产品代码行
3. 数据结束判定：在表头对应列中，若一行有 >=3 个空值（且已考虑合并单元格映射）并且不满足数据行特征，则判定当前产品线数据结束
"""

from __future__ import annotations

import argparse
import time
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont


EXPECTED_HEADERS = [
    "国家",
    "重量段/KG",
    "运费（RMB/KG）",
    "处理费(RMB/票)",
    "上网参考时效",
    "尺寸限制",
]

# 对表头做轻微归一化后匹配，兼容中英文括号/空格差异
HEADER_ALIASES = {
    "国家": "国家",
    "重量段kg": "重量段/KG",
    "重量段/kg": "重量段/KG",
    "重量段": "重量段/KG",
    "运费rmb/kg": "运费（RMB/KG）",
    "运费（rmb/kg）": "运费（RMB/KG）",
    "运费(rmb/kg)": "运费（RMB/KG）",
    "运费": "运费（RMB/KG）",
    "处理费rmb/票": "处理费(RMB/票)",
    "处理费（rmb/票）": "处理费(RMB/票)",
    "处理费(rmb/票)": "处理费(RMB/票)",
    "处理费": "处理费(RMB/票)",
    "上网参考时效": "上网参考时效",
    "尺寸限制": "尺寸限制",
}


@dataclass
class ProductLineTable:
    sheet_name: str
    product_line: str
    product_codes: List[str]
    update_time: str
    header_row: int
    data_start_row: int
    data_end_row: int
    headers: List[str]
    rows: List[Dict[str, object]]
    left_col: int
    right_col: int
    top_row: int
    bottom_row: int
    note_row: int
    note_text: str


class PricingExtractor:
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self._workbook = None

    @property
    def workbook(self):
        if self._workbook is None:
            self._workbook = load_workbook(self.excel_path, data_only=True)
        return self._workbook

    @staticmethod
    def _norm_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        text = text.replace("（", "(").replace("）", ")")
        text = re.sub(r"\s+", "", text)
        return text.lower()

    @staticmethod
    def _display_text(value: object) -> str:
        return "" if value is None else str(value).strip()

    def _build_merged_lookup(self, ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
        lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    lookup[(row, col)] = (min_row, min_col)
        return lookup

    def _get_value(self, ws: Worksheet, row: int, col: int, merged_lookup: Dict[Tuple[int, int], Tuple[int, int]]) -> object:
        value = ws.cell(row=row, column=col).value
        if value is not None:
            return value

        anchor = merged_lookup.get((row, col))
        if anchor:
            anchor_row, anchor_col = anchor
            return ws.cell(row=anchor_row, column=anchor_col).value
        return None

    def default_index_file(self) -> Path:
        return Path("output") / "index" / f"{self.excel_path.stem}_code_sheet_index.json"

    def build_code_sheet_index(self) -> Dict[str, Any]:
        code_to_sheets: Dict[str, set] = {}

        idx_wb = load_workbook(self.excel_path, data_only=True, read_only=True)
        try:
            letter_re = re.compile(r"[A-Za-z]")
            for ws in idx_wb.worksheets:
                print(f"[索引] 扫描工作表: {ws.title}", flush=True)
                for row in ws.iter_rows(values_only=True):
                    for value in row:
                        if value is None:
                            continue
                        text = str(value).strip()
                        if not text:
                            continue

                        # 快速匹配规则：包含“出口易”且包含英文字母
                        if "出口易" not in text or letter_re.search(text) is None:
                            continue

                        codes = self._extract_product_codes(text.upper())
                        for code in codes:
                            code_to_sheets.setdefault(code, set()).add(ws.title)
        finally:
            idx_wb.close()

        normalized = {k: sorted(list(v)) for k, v in sorted(code_to_sheets.items())}
        return {
            "excel": str(self.excel_path),
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "code_to_sheets": normalized,
        }

    @staticmethod
    def load_code_sheet_index(index_file: Path) -> Dict[str, Any]:
        return json.loads(index_file.read_text(encoding="utf-8"))

    @staticmethod
    def save_code_sheet_index(index_file: Path, index_data: Dict[str, Any]) -> None:
        index_file.parent.mkdir(parents=True, exist_ok=True)
        index_file.write_text(json.dumps(index_data, ensure_ascii=False, indent=2), encoding="utf-8")

    def get_sheets_by_code(self, code: str, index_file: Path, rebuild: bool = False) -> List[str]:
        if rebuild or not index_file.exists():
            print(f"[索引] 开始构建目录文件: {index_file}", flush=True)
            data = self.build_code_sheet_index()
            self.save_code_sheet_index(index_file, data)
            print(f"[索引] 目录构建完成: {index_file}", flush=True)
        else:
            print(f"[索引] 读取目录文件: {index_file}", flush=True)
            data = self.load_code_sheet_index(index_file)

        code_upper = code.strip().upper()
        mapping = data.get("code_to_sheets", {})
        sheets = mapping.get(code_upper, [])
        return sorted({s for s in sheets})

    def _match_header_row(self, ws: Worksheet, row: int, merged_lookup: Dict[Tuple[int, int], Tuple[int, int]]) -> Optional[Tuple[int, List[str]]]:
        max_col = ws.max_column
        # 6列表头，扫描起始列
        for start_col in range(1, max_col - 5 + 1):
            cells = [self._get_value(ws, row, start_col + i, merged_lookup) for i in range(6)]
            normalized = [self._norm_text(v) for v in cells]
            mapped: List[str] = []
            ok = True
            for item in normalized:
                if item not in HEADER_ALIASES:
                    ok = False
                    break
                mapped.append(HEADER_ALIASES[item])

            if ok and mapped == EXPECTED_HEADERS:
                return start_col, mapped
        return None

    @staticmethod
    def _extract_product_codes(text: str) -> List[str]:
        if not text:
            return []

        codes = set()
        for code in re.findall(r"[\(（]([A-Z]{2,8})[\)）]", text):
            codes.add(code)

        for code in re.findall(r"\b([A-Z]{2,8})\b", text):
            codes.add(code)

        return sorted(codes)

    @staticmethod
    def _looks_like_update_time(text: str) -> bool:
        if not text:
            return False
        if "更新" in text and ("年" in text or ":" in text or "：" in text):
            return True
        if re.search(r"20\d{2}[./-]?\d{1,2}[./-]?\d{1,2}", text):
            return True
        return False

    @staticmethod
    def _looks_like_weight(text: str) -> bool:
        if not text:
            return False
        return bool(re.search(r"[<>≤≥Ww]|kg|KG", text))

    @staticmethod
    def _looks_like_price(value: object) -> bool:
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        text = str(value).strip()
        return bool(re.fullmatch(r"\d+(\.\d+)?", text))

    @staticmethod
    def _is_blank(value: object) -> bool:
        if value is None:
            return True
        return str(value).strip() == ""

    @staticmethod
    def _looks_like_next_product_line(text: str) -> bool:
        if not text:
            return False
        # 常见产品线名称内会带代码括号，如 (CTD)
        if re.search(r"[\(（][A-Z]{2,8}[\)）]", text):
            return True
        # 排除普通数据行
        if "出口易" in text and ("USPS" in text or "Gofo" in text or "SPX" in text):
            return True
        return False

    @staticmethod
    def _looks_like_note_text(text: str) -> bool:
        if not text:
            return False
        key_words = ["备注", "说明", "注意", "注：", "注:", "特别说明", "服务说明"]
        if any(k in text for k in key_words):
            return True
        # 链接、地址、赔偿说明等也常作为备注出现
        if "http" in text.lower() or "赔偿" in text or "邮编" in text:
            return True
        return False

    def _find_note_row(
        self,
        ws: Worksheet,
        table: ProductLineTable,
        merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
    ) -> Tuple[int, str]:
        check_row = table.data_end_row + 1
        if check_row > ws.max_row:
            return 0, ""

        # 如果下一行是新表头，说明没有备注
        if self._match_header_row(ws, check_row, merged_lookup) is not None:
            return 0, ""

        values = [self._get_value(ws, check_row, table.left_col + i, merged_lookup) for i in range(6)]
        text_values = [self._display_text(v) for v in values if not self._is_blank(v)]
        dedup_text_values: List[str] = []
        seen = set()
        for item in text_values:
            if item not in seen:
                seen.add(item)
                dedup_text_values.append(item)
        row_text = " ".join(dedup_text_values)
        if not row_text:
            return 0, ""

        # 备注行一般是“合并大文本”或“非价格结构行”
        empty_count = sum(1 for v in values if self._is_blank(v))
        has_weight = self._looks_like_weight(self._display_text(values[1]))
        has_price = self._looks_like_price(values[2]) and self._looks_like_price(values[3])

        is_note_shape = empty_count >= 3 and not (has_weight or has_price)
        if is_note_shape or self._looks_like_note_text(row_text):
            return check_row, row_text

        return 0, ""

    def _extract_one_table(
        self,
        ws: Worksheet,
        header_row: int,
        start_col: int,
        headers: List[str],
        merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
    ) -> ProductLineTable:
        product_line_text = self._display_text(self._get_value(ws, header_row - 2, start_col, merged_lookup)) if header_row >= 3 else ""
        update_time_text = self._display_text(self._get_value(ws, header_row - 1, start_col, merged_lookup)) if header_row >= 2 else ""

        # 若标准位置为空，向该行其它列兜底找第一个非空文本
        if not product_line_text and header_row >= 3:
            for c in range(1, ws.max_column + 1):
                val = self._display_text(self._get_value(ws, header_row - 2, c, merged_lookup))
                if val:
                    product_line_text = val
                    break

        if not update_time_text and header_row >= 2:
            for c in range(1, ws.max_column + 1):
                val = self._display_text(self._get_value(ws, header_row - 1, c, merged_lookup))
                if val and self._looks_like_update_time(val):
                    update_time_text = val
                    break

        codes = self._extract_product_codes(product_line_text)

        rows: List[Dict[str, object]] = []
        data_start = header_row + 1
        data_end = header_row

        row = data_start
        while row <= ws.max_row:
            # 新表头出现，当前表结束
            maybe_next_header = self._match_header_row(ws, row, merged_lookup)
            if maybe_next_header is not None:
                break

            values = [self._get_value(ws, row, start_col + i, merged_lookup) for i in range(6)]
            text_values = [self._display_text(v) for v in values]

            if all(self._is_blank(v) for v in values):
                # 全空直接结束
                break

            empty_count = sum(1 for v in values if self._is_blank(v))
            country, weight, freight, handling, eta, size_limit = values

            has_country = not self._is_blank(country)
            has_weight = self._looks_like_weight(self._display_text(weight))
            has_price = self._looks_like_price(freight) and self._looks_like_price(handling)
            row_text = " ".join(t for t in text_values if t)

            # 命中“下一产品线”文本，结束当前数据
            if self._looks_like_next_product_line(row_text) and not has_weight:
                break

            # 结束规则：>=3 空，并且该行不具备数据行特征
            if empty_count >= 3 and not (has_weight or has_price or has_country):
                break

            # 只保留真正的数据行，防止把服务说明等混入
            if has_weight or (has_country and has_price):
                row_item = {
                    "国家": self._display_text(country),
                    "重量段/KG": self._display_text(weight),
                    "运费（RMB/KG）": freight,
                    "处理费(RMB/票)": handling,
                    "上网参考时效": self._display_text(eta),
                    "尺寸限制": self._display_text(size_limit),
                    "行号": row,
                }
                rows.append(row_item)
                data_end = row
            else:
                # 非数据行时，一般视作当前产品线数据已经结束
                break

            row += 1

        if not rows:
            data_start = header_row + 1
            data_end = header_row

        left_col = start_col
        right_col = start_col + 5
        top_row = max(1, header_row - 2)
        bottom_row = data_end

        table = ProductLineTable(
            sheet_name=ws.title,
            product_line=product_line_text,
            product_codes=codes,
            update_time=update_time_text,
            header_row=header_row,
            data_start_row=data_start,
            data_end_row=data_end,
            headers=headers,
            rows=rows,
            left_col=left_col,
            right_col=right_col,
            top_row=top_row,
            bottom_row=bottom_row,
            note_row=0,
            note_text="",
        )

        note_row, note_text = self._find_note_row(ws, table, merged_lookup)
        if note_row > 0:
            table.note_row = note_row
            table.note_text = note_text
            table.bottom_row = max(table.bottom_row, note_row)

        return table

    @staticmethod
    def _safe_filename(text: str) -> str:
        cleaned = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", text).strip(" ._")
        return cleaned or "unnamed"

    @staticmethod
    def _load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        font_candidates = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for font_path in font_candidates:
            p = Path(font_path)
            if p.exists():
                try:
                    return ImageFont.truetype(str(p), size)
                except OSError:
                    continue
        return ImageFont.load_default()

    def _render_table_snapshot(
        self,
        ws: Worksheet,
        table: ProductLineTable,
        merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
        output_path: Path,
    ) -> None:
        if table.bottom_row < table.top_row or table.right_col < table.left_col:
            return

        rows = list(range(table.top_row, table.bottom_row + 1))
        cols = list(range(table.left_col, table.right_col + 1))

        # 基于工作表行高列宽估算截图尺寸，尽量接近 Excel 可视效果
        col_px: Dict[int, int] = {}
        row_px: Dict[int, int] = {}
        for col in cols:
            letter = get_column_letter(col)
            width = ws.column_dimensions[letter].width
            if width is None:
                width = 12
            col_px[col] = max(70, int(float(width) * 8))

        for row in rows:
            height = ws.row_dimensions[row].height
            if height is None:
                height = 20
            row_px[row] = max(24, int(float(height) * 1.6))

        total_w = sum(col_px[c] for c in cols) + 2
        total_h = sum(row_px[r] for r in rows) + 2

        image = Image.new("RGB", (total_w, total_h), "white")
        draw = ImageDraw.Draw(image)
        font = self._load_font(14)

        # 绘制网格和文本
        y = 1
        for row in rows:
            x = 1
            for col in cols:
                cw = col_px[col]
                rh = row_px[row]
                draw.rectangle([x, y, x + cw, y + rh], outline=(0, 0, 0), width=1)

                value = self._get_value(ws, row, col, merged_lookup)
                text = self._display_text(value)
                if text:
                    text = text.replace("\n", " ")
                    draw.text((x + 4, y + 4), text, fill=(0, 0, 0), font=font)

                x += cw
            y += row_px[row]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(output_path)

    def _save_snapshots_via_excel_com(self, tables: List[ProductLineTable], output_dir: Path) -> List[Path]:
        try:
            import win32com.client  # type: ignore
            from PIL import ImageGrab  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "未安装 pywin32，无法使用 COM 渲染截图。请先安装: pip install pywin32"
            ) from exc

        output_dir.mkdir(parents=True, exist_ok=True)
        saved_files: List[Path] = []

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = None
        try:
            workbook = excel.Workbooks.Open(str(self.excel_path.resolve()))

            for idx, table in enumerate(tables, start=1):
                sheet = workbook.Worksheets(table.sheet_name)
                top_left = f"{get_column_letter(table.left_col)}{table.top_row}"
                bottom_right = f"{get_column_letter(table.right_col)}{table.bottom_row}"
                rng = sheet.Range(f"{top_left}:{bottom_right}")

                code_part = "_".join(table.product_codes) if table.product_codes else "NO_CODE"
                code_part = self._safe_filename(code_part)
                name_part = self._safe_filename(table.product_line[:30])
                file_name = f"{idx:03d}_{code_part}_{name_part}.png"
                file_path = output_dir / file_name

                # CopyPicture 会复制 Excel 原始视觉样式（字体/底色/边框）到剪贴板
                rng.CopyPicture(Appearance=1, Format=2)
                time.sleep(0.15)

                image = ImageGrab.grabclipboard()
                if image is None:
                    raise RuntimeError(
                        f"COM 截图失败: 无法从剪贴板获取图像，范围={top_left}:{bottom_right}"
                    )

                image.save(file_path)
                saved_files.append(file_path)

            return saved_files
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
            excel.Quit()

    def save_snapshots(self, tables: List[ProductLineTable], output_dir: Path) -> List[Path]:
        output_dir.mkdir(parents=True, exist_ok=True)
        saved_files: List[Path] = []

        for idx, table in enumerate(tables, start=1):
            ws = self.workbook[table.sheet_name]
            merged_lookup = self._build_merged_lookup(ws)

            code_part = "_".join(table.product_codes) if table.product_codes else "NO_CODE"
            code_part = self._safe_filename(code_part)
            name_part = self._safe_filename(table.product_line[:30])
            file_name = f"{idx:03d}_{code_part}_{name_part}.png"
            file_path = output_dir / file_name

            self._render_table_snapshot(ws, table, merged_lookup, file_path)
            saved_files.append(file_path)

        return saved_files

    def save_snapshots_com(self, tables: List[ProductLineTable], output_dir: Path) -> List[Path]:
        return self._save_snapshots_via_excel_com(tables, output_dir)

    def extract(self, include_sheets: Optional[List[str]] = None) -> List[ProductLineTable]:
        all_tables: List[ProductLineTable] = []
        target_sheets = include_sheets or self.workbook.sheetnames
        unknown = [s for s in target_sheets if s not in self.workbook.sheetnames]
        if unknown:
            raise ValueError(f"未找到工作表: {', '.join(unknown)}")

        for sheet_name in target_sheets:
            ws = self.workbook[sheet_name]
            merged_lookup = self._build_merged_lookup(ws)
            row = 1
            while row <= ws.max_row:
                header_info = self._match_header_row(ws, row, merged_lookup)
                if header_info is None:
                    row += 1
                    continue

                start_col, headers = header_info
                table = self._extract_one_table(ws, row, start_col, headers, merged_lookup)

                if table.rows:
                    all_tables.append(table)
                    row = max(table.data_end_row + 1, row + 1)
                else:
                    row += 1

        return all_tables


def summarize_by_sheet(tables: List[ProductLineTable]) -> Dict[str, Dict[str, object]]:
    summary: Dict[str, Dict[str, object]] = {}
    for table in tables:
        node = summary.setdefault(
            table.sheet_name,
            {
                "sheet": table.sheet_name,
                "table_count": 0,
                "row_count": 0,
                "product_codes": set(),
            },
        )
        node["table_count"] = int(node["table_count"]) + 1
        node["row_count"] = int(node["row_count"]) + len(table.rows)
        for c in table.product_codes:
            node["product_codes"].add(c)

    for node in summary.values():
        node["product_codes"] = sorted(list(node["product_codes"]))
        node["product_code_count"] = len(node["product_codes"])

    return summary


def tables_to_records(tables: List[ProductLineTable]) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for table in tables:
        for row in table.rows:
            record = {
                "sheet": table.sheet_name,
                "产品线": table.product_line,
                "产品代码": ",".join(table.product_codes),
                "更新时间": table.update_time,
                "header_row": table.header_row,
            }
            record.update(row)
            records.append(record)
    return records


def write_outputs(tables: List[ProductLineTable], output_json: Path, output_xlsx: Path) -> None:
    output_json.parent.mkdir(parents=True, exist_ok=True)

    payload = []
    for table in tables:
        payload.append(
            {
                "sheet": table.sheet_name,
                "产品线": table.product_line,
                "产品代码": table.product_codes,
                "更新时间": table.update_time,
                "header_row": table.header_row,
                "data_start_row": table.data_start_row,
                "data_end_row": table.data_end_row,
                "headers": table.headers,
                "range": {
                    "top_left": f"{get_column_letter(table.left_col)}{table.top_row}",
                    "bottom_right": f"{get_column_letter(table.right_col)}{table.bottom_row}",
                },
                "note": {
                    "has_note": table.note_row > 0,
                    "row": table.note_row,
                    "text": table.note_text,
                },
                "rows": table.rows,
            }
        )

    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # 延迟导入 pandas，避免作为核心解析依赖
    import pandas as pd  # pylint: disable=import-outside-toplevel

    df = pd.DataFrame(tables_to_records(tables))
    df.to_excel(output_xlsx, index=False)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="提取报价表中的每个产品线价格明细")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("出口易物流推广报价表2026.04.22.xlsx"),
        help="输入 Excel 文件路径",
    )
    parser.add_argument(
        "--sheets",
        type=str,
        default="",
        help="指定要查询的工作表，多个用逗号分隔；为空时扫描全部工作表",
    )
    parser.add_argument(
        "--code",
        type=str,
        default="",
        help="指定产品代码；会先通过产品代码-工作表目录进行匹配，只在匹配工作表执行",
    )
    parser.add_argument(
        "--index-file",
        type=Path,
        default=None,
        help="产品代码-工作表目录文件路径；为空时使用默认路径 output/index/<excel_stem>_code_sheet_index.json",
    )
    parser.add_argument(
        "--rebuild-index",
        action="store_true",
        help="强制重建产品代码-工作表目录文件",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("output") / "pricing_details.json",
        help="JSON 输出路径",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("output") / "pricing_details.xlsx",
        help="Excel 输出路径",
    )
    parser.add_argument(
        "--snapshot-dir",
        type=Path,
        default=Path("output") / "snapshots",
        help="产品代码价格信息截图输出目录",
    )
    parser.add_argument(
        "--snapshot-engine",
        choices=["draw", "com"],
        default="draw",
        help="截图引擎: draw=Python绘制, com=Excel COM原样渲染(Windows)",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()

    extractor = PricingExtractor(args.input)
    include_sheets: Optional[List[str]] = None

    if args.code.strip():
        index_file = args.index_file or extractor.default_index_file()
        matched_sheets = extractor.get_sheets_by_code(
            code=args.code,
            index_file=index_file,
            rebuild=args.rebuild_index,
        )

        if not matched_sheets:
            print(f"产品代码 {args.code.strip().upper()} 匹配不到工作表，已停止执行。")
            print(f"目录文件: {index_file}")
            return

        include_sheets = matched_sheets
        print(f"产品代码 {args.code.strip().upper()} 匹配工作表: {', '.join(include_sheets)}")
        print(f"目录文件: {index_file}")
    elif args.sheets.strip():
        include_sheets = [x.strip() for x in args.sheets.split(",") if x.strip()]

    tables = extractor.extract(include_sheets=include_sheets)

    write_outputs(tables, args.output_json, args.output_xlsx)
    if args.snapshot_engine == "com":
        snapshots = extractor.save_snapshots_com(tables, args.snapshot_dir)
    else:
        snapshots = extractor.save_snapshots(tables, args.snapshot_dir)

    print(f"识别到产品线数量: {len(tables)}")
    sheet_summary = summarize_by_sheet(tables)
    print(f"命中工作表数量: {len(sheet_summary)}")
    for sheet_name, info in sheet_summary.items():
        print(
            f"- {sheet_name}: 表格{info['table_count']}个, 数据行{info['row_count']}行, 产品代码{info['product_code_count']}个"
        )
    for idx, table in enumerate(tables, start=1):
        codes = ",".join(table.product_codes) if table.product_codes else "(无代码)"
        print(
            f"[{idx}] sheet={table.sheet_name} | 代码={codes} | 更新时间={table.update_time} | 数据行={len(table.rows)}"
        )

    print(f"JSON输出: {args.output_json}")
    print(f"Excel输出: {args.output_xlsx}")
    print(f"截图引擎: {args.snapshot_engine}")
    print(f"截图输出目录: {args.snapshot_dir}")
    print(f"截图数量: {len(snapshots)}")


if __name__ == "__main__":
    main()

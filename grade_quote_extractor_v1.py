#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
提取“等级报价”工作表中的产品代码分国家等级报价。

输出:
1) records: 扁平明细（每行一条 产品代码-国家分区-重量段）
2) by_product_code: 按产品代码聚合的结构化数据
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class GradeColumnMap:
    product_name_col: int
    service_code_col: int
    country_zone_col: int
    weight_col: int
    grade_effective_time_col: int
    d_freight_col: int
    d_handling_col: int
    c_freight_col: int
    c_handling_col: int
    b_freight_col: int
    b_handling_col: int
    a_freight_col: int
    a_handling_col: int
    external_freight_col: int
    external_handling_col: int
    external_effective_time_col: int


class GradeQuoteExtractor:
    def __init__(self, excel_path: Path, sheet_name: str = "等级报价"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.workbook = load_workbook(excel_path, data_only=True)
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"未找到工作表: {sheet_name}")
        self.ws: Worksheet = self.workbook[sheet_name]
        self.merged_lookup = self._build_merged_lookup(self.ws)

    @staticmethod
    def _build_merged_lookup(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
        lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    lookup[(r, c)] = (min_row, min_col)
        return lookup

    def _value(self, row: int, col: int) -> Any:
        val = self.ws.cell(row=row, column=col).value
        if val is not None:
            return val
        anchor = self.merged_lookup.get((row, col))
        if anchor:
            return self.ws.cell(row=anchor[0], column=anchor[1]).value
        return None

    @staticmethod
    def _text(val: Any) -> str:
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _norm(text: str) -> str:
        s = text.strip().lower()
        s = s.replace("（", "(").replace("）", ")")
        s = re.sub(r"\s+", "", s)
        return s

    @staticmethod
    def _is_blank(val: Any) -> bool:
        return val is None or str(val).strip() == ""

    @staticmethod
    def _to_number_or_text(val: Any) -> Any:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return val
        s = str(val).strip()
        if s == "":
            return None
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            try:
                if "." in s:
                    return float(s)
                return int(s)
            except ValueError:
                return s
        return s

    @staticmethod
    def _split_service_codes(service_text: str) -> List[str]:
        if not service_text:
            return []
        # 提取形如 SUX / SUT / SUR / USE 的代码，兼容 &,/,+,逗号、换行等连接符
        codes = re.findall(r"\b[A-Z][A-Z0-9]{1,8}\b", service_text.upper())
        seen = set()
        result: List[str] = []
        for code in codes:
            if code not in seen:
                seen.add(code)
                result.append(code)
        return result

    def _find_header_row(self) -> int:
        # 通常是第2行，但做扫描增强鲁棒性
        for r in range(1, min(self.ws.max_row, 20) + 1):
            vals = [self._norm(self._text(self._value(r, c))) for c in range(1, min(self.ws.max_column, 30) + 1)]
            row_text = "|".join(vals)
            if all(k in row_text for k in ["产品名称", "服务代码", "财务设置国家的分区", "重量段kg"]):
                return r
        raise ValueError("未识别到等级报价表头")

    def _build_column_map(self, header_row: int) -> GradeColumnMap:
        sub_row = header_row + 1
        max_col = self.ws.max_column

        product_name_col = service_code_col = country_zone_col = weight_col = grade_effective_time_col = -1
        d_freight_col = d_handling_col = -1
        c_freight_col = c_handling_col = -1
        b_freight_col = b_handling_col = -1
        a_freight_col = a_handling_col = -1
        external_freight_col = external_handling_col = external_effective_time_col = -1

        for c in range(1, max_col + 1):
            top = self._norm(self._text(self._value(header_row, c)))
            sub = self._norm(self._text(self._value(sub_row, c)))

            if top == "产品名称":
                product_name_col = c
            elif top == "服务代码":
                service_code_col = c
            elif top == "财务设置国家的分区":
                country_zone_col = c
            elif top == "重量段kg":
                weight_col = c
            elif top == "等级价格生效时间":
                grade_effective_time_col = c
            elif top == "d等级":
                if sub == "运费":
                    d_freight_col = c
                elif sub == "处理费":
                    d_handling_col = c
            elif top == "c等级":
                if sub == "运费":
                    c_freight_col = c
                elif sub == "处理费":
                    c_handling_col = c
            elif top == "b等级":
                if sub == "运费":
                    b_freight_col = c
                elif sub == "处理费":
                    b_handling_col = c
            elif top == "a等级":
                if sub == "运费":
                    a_freight_col = c
                elif sub == "处理费":
                    a_handling_col = c
            elif top == "对外报价":
                if sub == "运费":
                    external_freight_col = c
                elif sub == "处理费":
                    external_handling_col = c
                elif sub == "对外报价生效时间":
                    external_effective_time_col = c

        missing = {
            "产品名称": product_name_col,
            "服务代码": service_code_col,
            "财务设置国家的分区": country_zone_col,
            "重量段KG": weight_col,
            "等级价格生效时间": grade_effective_time_col,
            "D运费": d_freight_col,
            "D处理费": d_handling_col,
            "C运费": c_freight_col,
            "C处理费": c_handling_col,
            "B运费": b_freight_col,
            "B处理费": b_handling_col,
            "A运费": a_freight_col,
            "A处理费": a_handling_col,
            "对外运费": external_freight_col,
            "对外处理费": external_handling_col,
            "对外报价生效时间": external_effective_time_col,
        }
        miss = [k for k, v in missing.items() if v <= 0]
        if miss:
            raise ValueError(f"表头识别不完整，缺失列: {', '.join(miss)}")

        return GradeColumnMap(
            product_name_col=product_name_col,
            service_code_col=service_code_col,
            country_zone_col=country_zone_col,
            weight_col=weight_col,
            grade_effective_time_col=grade_effective_time_col,
            d_freight_col=d_freight_col,
            d_handling_col=d_handling_col,
            c_freight_col=c_freight_col,
            c_handling_col=c_handling_col,
            b_freight_col=b_freight_col,
            b_handling_col=b_handling_col,
            a_freight_col=a_freight_col,
            a_handling_col=a_handling_col,
            external_freight_col=external_freight_col,
            external_handling_col=external_handling_col,
            external_effective_time_col=external_effective_time_col,
        )

    def _is_valid_data_row(self, row: int, cm: GradeColumnMap) -> bool:
        weight = self._text(self._value(row, cm.weight_col))
        if not weight:
            return False
        has_digit = bool(re.search(r"\d", weight))
        if not has_digit:
            return False

        candidates = [
            self._value(row, cm.d_freight_col),
            self._value(row, cm.d_handling_col),
            self._value(row, cm.c_freight_col),
            self._value(row, cm.c_handling_col),
            self._value(row, cm.b_freight_col),
            self._value(row, cm.b_handling_col),
            self._value(row, cm.a_freight_col),
            self._value(row, cm.a_handling_col),
            self._value(row, cm.external_freight_col),
            self._value(row, cm.external_handling_col),
        ]
        return any(not self._is_blank(v) for v in candidates)

    def extract(self) -> Dict[str, Any]:
        header_row = self._find_header_row()
        cm = self._build_column_map(header_row)

        records: List[Dict[str, Any]] = []
        data_start = header_row + 2

        for r in range(data_start, self.ws.max_row + 1):
            if not self._is_valid_data_row(r, cm):
                continue

            product_name = self._text(self._value(r, cm.product_name_col))
            service_text = self._text(self._value(r, cm.service_code_col))
            country_zone = self._text(self._value(r, cm.country_zone_col))
            weight_range = self._text(self._value(r, cm.weight_col))
            grade_effective_time = self._text(self._value(r, cm.grade_effective_time_col))

            service_codes = self._split_service_codes(service_text)
            if not service_codes:
                continue

            base_payload = {
                "sheet": self.sheet_name,
                "row": r,
                "产品名称": product_name,
                "服务代码原文": service_text,
                "国家分区": country_zone,
                "重量段KG": weight_range,
                "等级价格生效时间": grade_effective_time,
                "D等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.d_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.d_handling_col)),
                },
                "C等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.c_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.c_handling_col)),
                },
                "B等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.b_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.b_handling_col)),
                },
                "A等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.a_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.a_handling_col)),
                },
                "对外报价": {
                    "运费": self._to_number_or_text(self._value(r, cm.external_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.external_handling_col)),
                    "对外报价生效时间": self._text(self._value(r, cm.external_effective_time_col)),
                },
            }

            for code in service_codes:
                row_item = dict(base_payload)
                row_item["产品代码"] = code
                records.append(row_item)

        by_product_code: Dict[str, Any] = {}
        for rec in records:
            code = rec["产品代码"]
            node = by_product_code.setdefault(
                code,
                {
                    "产品代码": code,
                    "产品名称": rec["产品名称"],
                    "服务代码原文样本": rec["服务代码原文"],
                    "国家分区报价": {},
                },
            )

            country = rec["国家分区"] or "(空国家分区)"
            node["国家分区报价"].setdefault(country, []).append(
                {
                    "重量段KG": rec["重量段KG"],
                    "等级价格生效时间": rec["等级价格生效时间"],
                    "D等级": rec["D等级"],
                    "C等级": rec["C等级"],
                    "B等级": rec["B等级"],
                    "A等级": rec["A等级"],
                    "对外报价": rec["对外报价"],
                    "来源行": rec["row"],
                }
            )

        return {
            "meta": {
                "excel": str(self.excel_path),
                "sheet": self.sheet_name,
                "header_row": header_row,
                "data_start_row": data_start,
                "record_count": len(records),
                "product_code_count": len(by_product_code),
            },
            "records": records,
            "by_product_code": by_product_code,
        }


def extract_grade_quotes(excel_path: str | Path, sheet_name: str = "等级报价") -> Dict[str, Any]:
    extractor = GradeQuoteExtractor(Path(excel_path), sheet_name=sheet_name)
    return extractor.extract()


def save_result(result: Dict[str, Any], output_json: Path, output_xlsx: Path) -> None:
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    # 扁平化输出到 Excel
    import pandas as pd  # pylint: disable=import-outside-toplevel

    rows: List[Dict[str, Any]] = []
    for rec in result["records"]:
        rows.append(
            {
                "sheet": rec["sheet"],
                "row": rec["row"],
                "产品名称": rec["产品名称"],
                "产品代码": rec["产品代码"],
                "服务代码原文": rec["服务代码原文"],
                "国家分区": rec["国家分区"],
                "重量段KG": rec["重量段KG"],
                "等级价格生效时间": rec["等级价格生效时间"],
                "D_运费": rec["D等级"]["运费"],
                "D_处理费": rec["D等级"]["处理费"],
                "C_运费": rec["C等级"]["运费"],
                "C_处理费": rec["C等级"]["处理费"],
                "B_运费": rec["B等级"]["运费"],
                "B_处理费": rec["B等级"]["处理费"],
                "A_运费": rec["A等级"]["运费"],
                "A_处理费": rec["A等级"]["处理费"],
                "对外_运费": rec["对外报价"]["运费"],
                "对外_处理费": rec["对外报价"]["处理费"],
                "对外报价生效时间": rec["对外报价"]["对外报价生效时间"],
            }
        )

    pd.DataFrame(rows).to_excel(output_xlsx, index=False)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="提取等级报价工作表中的产品代码分国家等级报价")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("2025年直发产品定价+vip 2026.04.22.xlsx"),
        help="输入Excel路径",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="等级报价",
        help="工作表名称",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("output") / "grade_quote_details.json",
        help="JSON输出路径",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("output") / "grade_quote_details.xlsx",
        help="Excel输出路径",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = extract_grade_quotes(args.input, sheet_name=args.sheet)
    save_result(result, args.output_json, args.output_xlsx)

    meta = result["meta"]
    print(f"工作表: {meta['sheet']}")
    print(f"表头行: {meta['header_row']}")
    print(f"数据起始行: {meta['data_start_row']}")
    print(f"明细记录数: {meta['record_count']}")
    print(f"产品代码数: {meta['product_code_count']}")
    print(f"JSON输出: {args.output_json}")
    print(f"Excel输出: {args.output_xlsx}")


if __name__ == "__main__":
    main()

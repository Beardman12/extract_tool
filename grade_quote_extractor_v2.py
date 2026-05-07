#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
提取“等级报价”工作表中的产品代码分国家等级报价。

输出:
1) records: 扁平明细（每行一条 产品代码-国家分区-重量段）
2) by_product_code: 按产品代码聚合的结构化数据
"""

from __future__ import annotations

import argparse
import time
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont


@dataclass
class GradeColumnMap:
    product_name_col: int
    service_code_col: int
    country_zone_col: int
    weight_col: int
    grade_effective_time_col: int
    d_freight_col: int
    d_handling_col: int
    c_freight_col: int
    c_handling_col: int
    b_freight_col: int
    b_handling_col: int
    a_freight_col: int
    a_handling_col: int
    external_freight_col: int
    external_handling_col: int
    external_effective_time_col: int


@dataclass
class ProductCodeBlock:
    product_name: str
    service_text: str
    service_codes: List[str]
    start_row: int
    end_row: int


class GradeQuoteExtractor:
    def __init__(self, excel_path: Path, sheet_name: str = "等级报价"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.workbook = load_workbook(excel_path, data_only=True)
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"未找到工作表: {sheet_name}")
        self.ws: Worksheet = self.workbook[sheet_name]
        self.merged_lookup = self._build_merged_lookup(self.ws)

    @staticmethod
    def _build_merged_lookup(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
        lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    lookup[(r, c)] = (min_row, min_col)
        return lookup

    def _value(self, row: int, col: int) -> Any:
        val = self.ws.cell(row=row, column=col).value
        if val is not None:
            return val
        anchor = self.merged_lookup.get((row, col))
        if anchor:
            return self.ws.cell(row=anchor[0], column=anchor[1]).value
        return None

    @staticmethod
    def _text(val: Any) -> str:
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _norm(text: str) -> str:
        s = text.strip().lower()
        s = s.replace("（", "(").replace("）", ")")
        s = re.sub(r"\s+", "", s)
        return s

    @staticmethod
    def _is_blank(val: Any) -> bool:
        return val is None or str(val).strip() == ""

    @staticmethod
    def _to_number_or_text(val: Any) -> Any:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return val
        s = str(val).strip()
        if s == "":
            return None
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            try:
                if "." in s:
                    return float(s)
                return int(s)
            except ValueError:
                return s
        return s

    @staticmethod
    def _split_service_codes(service_text: str) -> List[str]:
        if not service_text:
            return []
        # 提取形如 SUX / SUT / SUR / USE 的代码，兼容 &,/,+,逗号、换行等连接符
        codes = re.findall(r"\b[A-Z][A-Z0-9]{1,8}\b", service_text.upper())
        seen = set()
        result: List[str] = []
        for code in codes:
            if code not in seen:
                seen.add(code)
                result.append(code)
        return result

    @staticmethod
    def _safe_filename(text: str) -> str:
        cleaned = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", text).strip(" ._")
        return cleaned or "unnamed"

    @staticmethod
    def _load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        candidates = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for p in candidates:
            path = Path(p)
            if path.exists():
                try:
                    return ImageFont.truetype(str(path), size)
                except OSError:
                    continue
        return ImageFont.load_default()

    def _find_header_row(self) -> int:
        # 通常是第2行，但做扫描增强鲁棒性
        for r in range(1, min(self.ws.max_row, 20) + 1):
            vals = [self._norm(self._text(self._value(r, c))) for c in range(1, min(self.ws.max_column, 30) + 1)]
            row_text = "|".join(vals)
            if all(k in row_text for k in ["产品名称", "服务代码", "财务设置国家的分区", "重量段kg"]):
                return r
        raise ValueError("未识别到等级报价表头")

    def _build_column_map(self, header_row: int) -> GradeColumnMap:
        sub_row = header_row + 1
        max_col = self.ws.max_column

        product_name_col = service_code_col = country_zone_col = weight_col = grade_effective_time_col = -1
        d_freight_col = d_handling_col = -1
        c_freight_col = c_handling_col = -1
        b_freight_col = b_handling_col = -1
        a_freight_col = a_handling_col = -1
        external_freight_col = external_handling_col = external_effective_time_col = -1

        for c in range(1, max_col + 1):
            top = self._norm(self._text(self._value(header_row, c)))
            sub = self._norm(self._text(self._value(sub_row, c)))

            if top == "产品名称":
                product_name_col = c
            elif top == "服务代码":
                service_code_col = c
            elif top == "财务设置国家的分区":
                country_zone_col = c
            elif top == "重量段kg":
                weight_col = c
            elif top == "等级价格生效时间":
                grade_effective_time_col = c
            elif top == "d等级":
                if sub == "运费":
                    d_freight_col = c
                elif sub == "处理费":
                    d_handling_col = c
            elif top == "c等级":
                if sub == "运费":
                    c_freight_col = c
                elif sub == "处理费":
                    c_handling_col = c
            elif top == "b等级":
                if sub == "运费":
                    b_freight_col = c
                elif sub == "处理费":
                    b_handling_col = c
            elif top == "a等级":
                if sub == "运费":
                    a_freight_col = c
                elif sub == "处理费":
                    a_handling_col = c
            elif top == "对外报价":
                if sub == "运费":
                    external_freight_col = c
                elif sub == "处理费":
                    external_handling_col = c
                elif sub == "对外报价生效时间":
                    external_effective_time_col = c

        missing = {
            "产品名称": product_name_col,
            "服务代码": service_code_col,
            "财务设置国家的分区": country_zone_col,
            "重量段KG": weight_col,
            "等级价格生效时间": grade_effective_time_col,
            "D运费": d_freight_col,
            "D处理费": d_handling_col,
            "C运费": c_freight_col,
            "C处理费": c_handling_col,
            "B运费": b_freight_col,
            "B处理费": b_handling_col,
            "A运费": a_freight_col,
            "A处理费": a_handling_col,
            "对外运费": external_freight_col,
            "对外处理费": external_handling_col,
            "对外报价生效时间": external_effective_time_col,
        }
        miss = [k for k, v in missing.items() if v <= 0]
        if miss:
            raise ValueError(f"表头识别不完整，缺失列: {', '.join(miss)}")

        return GradeColumnMap(
            product_name_col=product_name_col,
            service_code_col=service_code_col,
            country_zone_col=country_zone_col,
            weight_col=weight_col,
            grade_effective_time_col=grade_effective_time_col,
            d_freight_col=d_freight_col,
            d_handling_col=d_handling_col,
            c_freight_col=c_freight_col,
            c_handling_col=c_handling_col,
            b_freight_col=b_freight_col,
            b_handling_col=b_handling_col,
            a_freight_col=a_freight_col,
            a_handling_col=a_handling_col,
            external_freight_col=external_freight_col,
            external_handling_col=external_handling_col,
            external_effective_time_col=external_effective_time_col,
        )

    def _is_valid_data_row(self, row: int, cm: GradeColumnMap) -> bool:
        weight = self._text(self._value(row, cm.weight_col))
        if not weight:
            return False
        has_digit = bool(re.search(r"\d", weight))
        if not has_digit:
            return False

        candidates = [
            self._value(row, cm.d_freight_col),
            self._value(row, cm.d_handling_col),
            self._value(row, cm.c_freight_col),
            self._value(row, cm.c_handling_col),
            self._value(row, cm.b_freight_col),
            self._value(row, cm.b_handling_col),
            self._value(row, cm.a_freight_col),
            self._value(row, cm.a_handling_col),
            self._value(row, cm.external_freight_col),
            self._value(row, cm.external_handling_col),
        ]
        return any(not self._is_blank(v) for v in candidates)

    def _build_product_blocks(self, cm: GradeColumnMap, data_start: int) -> List[ProductCodeBlock]:
        blocks: List[ProductCodeBlock] = []
        current: Optional[ProductCodeBlock] = None

        for r in range(data_start, self.ws.max_row + 1):
            if not self._is_valid_data_row(r, cm):
                continue

            raw_service = self.ws.cell(row=r, column=cm.service_code_col).value
            service_text = self._text(self._value(r, cm.service_code_col))
            product_name = self._text(self._value(r, cm.product_name_col))
            codes = self._split_service_codes(service_text)

            # 原始服务代码列有值，视作新产品代码块的起点
            if raw_service is not None or current is None:
                if current is not None:
                    blocks.append(current)

                current = ProductCodeBlock(
                    product_name=product_name,
                    service_text=service_text,
                    service_codes=codes,
                    start_row=r,
                    end_row=r,
                )
            else:
                current.end_row = r

        if current is not None:
            blocks.append(current)

        return blocks

    def _render_snapshot_draw(
        self,
        top_row: int,
        bottom_row: int,
        left_col: int,
        right_col: int,
        output_path: Path,
    ) -> None:
        rows = list(range(top_row, bottom_row + 1))
        cols = list(range(left_col, right_col + 1))

        col_px: Dict[int, int] = {}
        row_px: Dict[int, int] = {}
        for c in cols:
            letter = get_column_letter(c)
            width = self.ws.column_dimensions[letter].width
            if width is None:
                width = 12
            col_px[c] = max(70, int(float(width) * 8))

        for r in rows:
            height = self.ws.row_dimensions[r].height
            if height is None:
                height = 20
            row_px[r] = max(24, int(float(height) * 1.6))

        total_w = sum(col_px[c] for c in cols) + 2
        total_h = sum(row_px[r] for r in rows) + 2

        image = Image.new("RGB", (total_w, total_h), "white")
        draw = ImageDraw.Draw(image)
        font = self._load_font(14)

        y = 1
        for r in rows:
            x = 1
            for c in cols:
                cw = col_px[c]
                rh = row_px[r]
                draw.rectangle([x, y, x + cw, y + rh], outline=(0, 0, 0), width=1)
                txt = self._text(self._value(r, c)).replace("\n", " ")
                if txt:
                    draw.text((x + 4, y + 4), txt, fill=(0, 0, 0), font=font)
                x += cw
            y += row_px[r]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(output_path)

    def _render_snapshot_com(
        self,
        top_row: int,
        bottom_row: int,
        left_col: int,
        right_col: int,
        output_path: Path,
    ) -> None:
        import win32com.client  # type: ignore
        from PIL import ImageGrab  # type: ignore

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = None
        try:
            wb = excel.Workbooks.Open(str(self.excel_path.resolve()))
            ws = wb.Worksheets(self.sheet_name)

            top_left = f"{get_column_letter(left_col)}{top_row}"
            bottom_right = f"{get_column_letter(right_col)}{bottom_row}"
            ws.Range(f"{top_left}:{bottom_right}").CopyPicture(Appearance=1, Format=2)
            time.sleep(0.15)

            img = ImageGrab.grabclipboard()
            if img is None:
                raise RuntimeError(f"COM截图失败: {top_left}:{bottom_right}")

            output_path.parent.mkdir(parents=True, exist_ok=True)
            img.save(output_path)
        finally:
            if wb is not None:
                wb.Close(SaveChanges=False)
            excel.Quit()

    def save_block_snapshots(
        self,
        blocks: List[ProductCodeBlock],
        header_row: int,
        cm: GradeColumnMap,
        output_dir: Path,
        engine: str = "com",
    ) -> List[Dict[str, Any]]:
        output_dir.mkdir(parents=True, exist_ok=True)
        snapshots: List[Dict[str, Any]] = []

        top_row = header_row
        left_col = cm.product_name_col
        right_col = cm.external_effective_time_col

        for idx, block in enumerate(blocks, start=1):
            if not block.service_codes:
                continue

            for code in block.service_codes:
                code_part = self._safe_filename(code)
                name_part = self._safe_filename(block.product_name[:20])
                file_name = f"{idx:03d}_{code_part}_{name_part}_r{block.start_row}-{block.end_row}.png"
                file_path = output_dir / file_name

                used_engine = engine
                try:
                    if engine == "com":
                        self._render_snapshot_com(top_row, block.end_row, left_col, right_col, file_path)
                    else:
                        self._render_snapshot_draw(top_row, block.end_row, left_col, right_col, file_path)
                except Exception:
                    # COM不可用时自动回退，确保有截图产出
                    used_engine = "draw"
                    self._render_snapshot_draw(top_row, block.end_row, left_col, right_col, file_path)

                snapshots.append(
                    {
                        "产品代码": code,
                        "产品名称": block.product_name,
                        "服务代码原文": block.service_text,
                        "row_range": f"{block.start_row}-{block.end_row}",
                        "range": {
                            "top_left": f"{get_column_letter(left_col)}{top_row}",
                            "bottom_right": f"{get_column_letter(right_col)}{block.end_row}",
                        },
                        "snapshot": str(file_path),
                        "engine": used_engine,
                    }
                )

        return snapshots

    def extract(self) -> Dict[str, Any]:
        header_row = self._find_header_row()
        cm = self._build_column_map(header_row)

        records: List[Dict[str, Any]] = []
        data_start = header_row + 2

        for r in range(data_start, self.ws.max_row + 1):
            if not self._is_valid_data_row(r, cm):
                continue

            product_name = self._text(self._value(r, cm.product_name_col))
            service_text = self._text(self._value(r, cm.service_code_col))
            country_zone = self._text(self._value(r, cm.country_zone_col))
            weight_range = self._text(self._value(r, cm.weight_col))
            grade_effective_time = self._text(self._value(r, cm.grade_effective_time_col))

            service_codes = self._split_service_codes(service_text)
            if not service_codes:
                continue

            base_payload = {
                "sheet": self.sheet_name,
                "row": r,
                "产品名称": product_name,
                "服务代码原文": service_text,
                "国家分区": country_zone,
                "重量段KG": weight_range,
                "等级价格生效时间": grade_effective_time,
                "D等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.d_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.d_handling_col)),
                },
                "C等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.c_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.c_handling_col)),
                },
                "B等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.b_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.b_handling_col)),
                },
                "A等级": {
                    "运费": self._to_number_or_text(self._value(r, cm.a_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.a_handling_col)),
                },
                "对外报价": {
                    "运费": self._to_number_or_text(self._value(r, cm.external_freight_col)),
                    "处理费": self._to_number_or_text(self._value(r, cm.external_handling_col)),
                    "对外报价生效时间": self._text(self._value(r, cm.external_effective_time_col)),
                },
            }

            for code in service_codes:
                row_item = dict(base_payload)
                row_item["产品代码"] = code
                records.append(row_item)

        by_product_code: Dict[str, Any] = {}
        for rec in records:
            code = rec["产品代码"]
            node = by_product_code.setdefault(
                code,
                {
                    "产品代码": code,
                    "产品名称": rec["产品名称"],
                    "服务代码原文样本": rec["服务代码原文"],
                    "国家分区报价": {},
                },
            )

            country = rec["国家分区"] or "(空国家分区)"
            node["国家分区报价"].setdefault(country, []).append(
                {
                    "重量段KG": rec["重量段KG"],
                    "等级价格生效时间": rec["等级价格生效时间"],
                    "D等级": rec["D等级"],
                    "C等级": rec["C等级"],
                    "B等级": rec["B等级"],
                    "A等级": rec["A等级"],
                    "对外报价": rec["对外报价"],
                    "来源行": rec["row"],
                }
            )

        blocks = self._build_product_blocks(cm, data_start)

        return {
            "meta": {
                "excel": str(self.excel_path),
                "sheet": self.sheet_name,
                "header_row": header_row,
                "data_start_row": data_start,
                "record_count": len(records),
                "product_code_count": len(by_product_code),
                "block_count": len(blocks),
            },
            "records": records,
            "by_product_code": by_product_code,
            "blocks": [
                {
                    "产品名称": b.product_name,
                    "服务代码原文": b.service_text,
                    "产品代码列表": b.service_codes,
                    "start_row": b.start_row,
                    "end_row": b.end_row,
                }
                for b in blocks
            ],
        }


def extract_grade_quotes(excel_path: str | Path, sheet_name: str = "等级报价") -> Dict[str, Any]:
    extractor = GradeQuoteExtractor(Path(excel_path), sheet_name=sheet_name)
    return extractor.extract()


def save_result(result: Dict[str, Any], output_json: Path, output_xlsx: Path) -> None:
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    # 扁平化输出到 Excel
    import pandas as pd  # pylint: disable=import-outside-toplevel

    rows: List[Dict[str, Any]] = []
    for rec in result["records"]:
        rows.append(
            {
                "sheet": rec["sheet"],
                "row": rec["row"],
                "产品名称": rec["产品名称"],
                "产品代码": rec["产品代码"],
                "服务代码原文": rec["服务代码原文"],
                "国家分区": rec["国家分区"],
                "重量段KG": rec["重量段KG"],
                "等级价格生效时间": rec["等级价格生效时间"],
                "D_运费": rec["D等级"]["运费"],
                "D_处理费": rec["D等级"]["处理费"],
                "C_运费": rec["C等级"]["运费"],
                "C_处理费": rec["C等级"]["处理费"],
                "B_运费": rec["B等级"]["运费"],
                "B_处理费": rec["B等级"]["处理费"],
                "A_运费": rec["A等级"]["运费"],
                "A_处理费": rec["A等级"]["处理费"],
                "对外_运费": rec["对外报价"]["运费"],
                "对外_处理费": rec["对外报价"]["处理费"],
                "对外报价生效时间": rec["对外报价"]["对外报价生效时间"],
            }
        )

    pd.DataFrame(rows).to_excel(output_xlsx, index=False)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="提取等级报价工作表中的产品代码分国家等级报价")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("2025年直发产品定价+vip 2026.04.22.xlsx"),
        help="输入Excel路径",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="等级报价",
        help="工作表名称",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("output") / "grade_quote_details.json",
        help="JSON输出路径",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("output") / "grade_quote_details.xlsx",
        help="Excel输出路径",
    )
    parser.add_argument(
        "--snapshot-dir",
        type=Path,
        default=Path("output") / "grade_quote_snapshots",
        help="产品代码截图输出目录",
    )
    parser.add_argument(
        "--snapshot-engine",
        choices=["com", "draw"],
        default="com",
        help="截图引擎: com=Excel原样截图, draw=程序绘制",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    extractor = GradeQuoteExtractor(Path(args.input), sheet_name=args.sheet)
    result = extractor.extract()

    header_row = int(result["meta"]["header_row"])
    cm = extractor._build_column_map(header_row)
    blocks: List[ProductCodeBlock] = []
    for b in result.get("blocks", []):
        blocks.append(
            ProductCodeBlock(
                product_name=b["产品名称"],
                service_text=b["服务代码原文"],
                service_codes=list(b["产品代码列表"]),
                start_row=int(b["start_row"]),
                end_row=int(b["end_row"]),
            )
        )

    snapshots = extractor.save_block_snapshots(
        blocks=blocks,
        header_row=header_row,
        cm=cm,
        output_dir=args.snapshot_dir,
        engine=args.snapshot_engine,
    )
    result["snapshots"] = snapshots
    result["meta"]["snapshot_count"] = len(snapshots)

    save_result(result, args.output_json, args.output_xlsx)

    meta = result["meta"]
    print(f"工作表: {meta['sheet']}")
    print(f"表头行: {meta['header_row']}")
    print(f"数据起始行: {meta['data_start_row']}")
    print(f"明细记录数: {meta['record_count']}")
    print(f"产品代码数: {meta['product_code_count']}")
    print(f"产品块数量: {meta['block_count']}")
    print(f"截图数量: {meta['snapshot_count']}")
    print(f"截图引擎: {args.snapshot_engine}")
    print(f"截图目录: {args.snapshot_dir}")
    print(f"JSON输出: {args.output_json}")
    print(f"Excel输出: {args.output_xlsx}")


if __name__ == "__main__":
    main()

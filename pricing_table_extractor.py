#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
出口易物流报价表智能识别工具 v2.0
功能：
1. 自动扫描所有工作表，根据表头特征识别数据区域
2. 通过产品代码(如CTD、GFD)筛选对应的报价表
3. 支持修改重量段、运费、处理费后重新生成对比图片
4. 原内容和修改后的内容分别输出为图片
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from PIL import Image, ImageDraw, ImageFont
import os
import re

# 定义样式常量
BLACK = (0, 0, 0)
RED = (255, 0, 0)
WHITE = (255, 255, 255)
LIGHT_GRAY = (240, 240, 240)

# 中文字体路径
FONT_PATHS = [
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/System/Library/Fonts/PingFang.ttc",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
]

def get_font(size=12, bold=False):
    """获取可用的中文字体"""
    for path in FONT_PATHS:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except:
                continue
    return ImageFont.load_default()


class PricingTableParser:
    """
    物流报价表智能解析器
    通过表头特征自动识别数据区域
    """

    # 标准表头关键词
    HEADER_KEYWORDS = ['国家', '重量段', '运费', '处理费', '上网参考时效', '尺寸限制']

    def __init__(self, excel_path):
        self.wb = load_workbook(excel_path, data_only=True)
        self.tables = {}  # 存储所有解析出的报价表
        self._parse_all_sheets()

    def _find_header_row(self, ws, start_row, end_row):
        """在指定范围内查找表头行"""
        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(1, 15):  # 扫描足够多的列
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    row_values.append(str(val).strip())
                else:
                    row_values.append('')

            # 检查是否包含所有关键表头词
            row_text = ''.join(row_values)
            match_count = sum(1 for kw in self.HEADER_KEYWORDS if kw in row_text)

            if match_count >= 5:  # 至少匹配5个关键词
                # 确定表头的起始列
                header_start_col = 1
                for col_idx in range(1, 15):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val and str(val).strip() in self.HEADER_KEYWORDS:
                        header_start_col = col_idx
                        break
                return row_idx, header_start_col
        return None, None

    def _is_product_code(self, value):
        """判断是否产品代码（如CTD、GFD）"""
        if not value:
            return False
        value_str = str(value).strip()
        # 常见格式：单独大写字母组合、括号内的大写字母组合
        if re.match(r'^[A-Z]{2,6}$', value_str):
            return True
        if re.match(r'.*\(([A-Z]{2,6})\).*', value_str):
            return True
        return False

    def _is_data_row(self, row_values):
        """判断是否数据行"""
        if not row_values or len(row_values) < 3:
            return False
        # 检查是否有重量段或国家信息
        has_weight = any('W' in str(v) or 'KG' in str(v) or 'kg' in str(v) for v in row_values if v)
        has_country = any('国' in str(v) or '美' in str(v) or '美' in str(v) for v in row_values if v)
        return has_weight or (has_country and any(v for v in row_values if v))

    def _parse_table_block(self, ws, start_row, end_row, code_filter=None):
        """解析一个表格块"""
        tables = []

        current_row = start_row
        while current_row <= end_row:
            row_values = [ws.cell(row=current_row, column=col).value for col in range(1, 15)]
            row_text = ' '.join([str(v) if v else '' for v in row_values])

            # 跳过空行
            if not any(row_values):
                current_row += 1
                continue

            # 查找表头
            header_row, header_col = self._find_header_row(ws, current_row, min(current_row + 5, end_row))

            if header_row:
                # 向上查找更新时间和产品名称（通常在表头上方1-3行）
                update_time = None
                product_name = None
                product_codes = []

                for look_row in range(header_row - 1, max(start_row, header_row - 5), -1):
                    look_values = [ws.cell(row=look_row, column=col).value for col in range(1, 15)]
                    look_text = ''.join([str(v) if v else '' for v in look_values])

                    if '更新' in look_text or '时间' in look_text:
                        for v in look_values:
                            if v and ('20' in str(v) or '年' in str(v) or '月' in str(v)):
                                update_time = str(v)
                                break

                    if not product_name:
                        for v in look_values:
                            if v and len(str(v).strip()) > 3:
                                product_name = str(v).strip()
                                break

                    # 提取产品代码
                    for v in look_values:
                        if v:
                            # 检查括号内的代码
                            codes = re.findall(r'\(([A-Z]{2,6})\)', str(v))
                            product_codes.extend(codes)
                            # 检查独立的大写字母组合
                            independent_codes = re.findall(r'\b([A-Z]{3,6})\b', str(v))
                            for code in independent_codes:
                                if code not in product_codes and len(code) >= 3:
                                    product_codes.append(code)

                # 向下解析数据行
                data_rows = []
                for data_row in range(header_row + 1, end_row + 1):
                    data_values = [ws.cell(row=data_row, column=col).value
                                   for col in range(header_col, header_col + 8)]

                    # 检查是否结束（空行或新的表头）
                    if not any(data_values):
                        break

                    # 检查是否出现新的产品线（通过产品名称判断）
                    row_text = ' '.join([str(v) if v else '' for v in data_values])
                    new_header, _ = self._find_header_row(ws, data_row, data_row + 1)
                    if new_header and new_header != header_row:
                        break

                    # 如果有数据且不全是空值
                    if any(v for v in data_values if v):
                        data_rows.append(data_values)

                # 如果有数据且匹配代码过滤
                if data_rows and (not code_filter or code_filter in product_codes):
                    tables.append({
                        'product_name': product_name or '',
                        'product_codes': list(set(product_codes)),
                        'update_time': update_time or '更新时间未知',
                        'headers': data_values[:6] if data_rows else [],
                        'data_rows': data_rows,
                        'sheet_name': ws.title,
                        'header_row': header_row
                    })

                current_row = header_row + len(data_rows) + 1
            else:
                current_row += 1

        return tables

    def _parse_all_sheets(self):
        """解析所有工作表"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 扫描整个工作表寻找表头
            for row_idx in range(1, ws.max_row + 1):
                header_row, header_col = self._find_header_row(ws, row_idx, row_idx)

                if header_row:
                    # 向上查找产品信息
                    update_time = None
                    product_name = None
                    product_codes = []

                    for look_row in range(header_row - 1, max(1, header_row - 8), -1):
                        look_values = [ws.cell(row=look_row, column=col).value for col in range(1, 15)]
                        look_text = ''.join([str(v) if v else '' for v in look_values])

                        if '更新' in look_text:
                            for v in look_values:
                                if v and ('20' in str(v) or '年' in str(v)):
                                    update_time = str(v)
                                    break

                        if not product_name:
                            for v in look_values:
                                if v and len(str(v).strip()) > 5:
                                    product_name = str(v).strip()
                                    break

                        # 提取产品代码
                        for v in look_values:
                            if v:
                                codes = re.findall(r'\(([A-Z]{2,6})\)', str(v))
                                product_codes.extend(codes)

                    # 向下解析数据行
                    data_rows = []
                    for data_row in range(header_row + 1, min(ws.max_row + 1, header_row + 50)):
                        data_values = [ws.cell(row=data_row, column=col).value
                                       for col in range(header_col, header_col + 8)]

                        # 检查是否结束
                        if not any(data_values):
                            # 空一行继续检查
                            next_values = [ws.cell(row=data_row+1, column=col).value
                                          for col in range(header_col, header_col + 8)]
                            if not any(next_values):
                                break
                            continue

                        # 检查是否出现新的表头
                        if self._find_header_row(ws, data_row, data_row + 1)[0]:
                            break

                        if any(v for v in data_values if v):
                            data_rows.append(data_values)

                    # 存储表格
                    if data_rows:
                        table_info = {
                            'product_name': product_name or '',
                            'product_codes': list(set(product_codes)),
                            'update_time': update_time or '',
                            'headers': data_values[:6],
                            'data_rows': data_rows,
                            'sheet_name': sheet_name,
                            'header_row': header_row
                        }

                        # 按代码索引
                        for code in table_info['product_codes']:
                            self.tables[code] = table_info

                        # 也按产品名索引
                        if product_name:
                            self.tables[f"NAME_{product_name}"] = table_info

    def get_table(self, code):
        """获取指定代码的报价表"""
        return self.tables.get(code)

    def get_all_codes(self):
        """获取所有已识别的产品代码"""
        return [k for k in self.tables.keys() if not k.startswith('NAME_')]

    def list_all_tables(self):
        """列出所有已识别的报价表"""
        result = []
        seen = set()
        for key, table in self.tables.items():
            if key.startswith('NAME_'):
                continue
            if table['product_name'] not in seen:
                result.append({
                    'codes': table['product_codes'],
                    'name': table['product_name'],
                    'sheet': table['sheet_name']
                })
                seen.add(table['product_name'])
        return result


def create_table_image(data, output_path):
    """生成表格图片 - 仿Excel样式"""

    font_normal = get_font(14, False)
    font_bold = get_font(14, True)
    font_title = get_font(16, True)
    font_small = get_font(11, False)
    font_red = get_font(14, True)

    # 列宽设置
    col_widths = [100, 180, 140, 140, 120, 320]

    # 计算行高
    row_heights = [40, 25, 35]  # 标题、更新时间、表头

    # 数据行高
    for row_data in data['data_rows']:
        row_heights.append(35)

    # 尺寸限制行高
    row_heights.append(60)

    total_width = sum(col_widths) + 40
    total_height = sum(row_heights) + 20

    # 创建图片
    img = Image.new('RGB', (total_width, total_height), WHITE)
    draw = ImageDraw.Draw(img)

    x_offset = 20
    y_offset = 10

    # 绘制标题
    draw.rectangle([x_offset, y_offset, x_offset + total_width - 40, y_offset + row_heights[0]],
                   fill=WHITE, outline=BLACK, width=2)

    # 生成标题文本
    title_parts = []
    for code in data['product_codes'][:2]:
        title_parts.append(code)

    if len(data['product_codes']) > 2:
        title_text = f"{' / '.join(title_parts)} 等"
    else:
        title_text = ' / '.join(title_parts) if title_parts else (data['product_name'][:20] if data['product_name'] else '报价表')

    bbox = draw.textbbox((0, 0), title_text, font=font_title)
    text_width = bbox[2] - bbox[0]
    draw.text((x_offset + (total_width - 40 - text_width) // 2, y_offset + 12),
              title_text, fill=BLACK, font=font_title)

    y_offset += row_heights[0]

    # 绘制更新时间（红色）
    draw.rectangle([x_offset, y_offset, x_offset + total_width - 40, y_offset + row_heights[1]],
                   fill=WHITE, outline=BLACK, width=1)

    update_text = str(data['update_time'] or '更新时间：未知')
    bbox = draw.textbbox((0, 0), update_text, font=font_small)
    text_width = bbox[2] - bbox[0]
    draw.text((x_offset + total_width - 40 - text_width - 10, y_offset + 5),
              update_text, fill=RED, font=font_small)
    y_offset += row_heights[1]

    # 列位置
    col_positions = [x_offset]
    for w in col_widths[:-1]:
        col_positions.append(col_positions[-1] + w)

    # 绘制表头
    headers = data.get('headers', ['国家', '重量段/KG', '运费(RMB/KG)', '处理费(RMB/票)', '上网参考时效', '尺寸限制'])

    draw.rectangle([x_offset, y_offset, x_offset + sum(col_widths), y_offset + row_heights[2]],
                   fill=LIGHT_GRAY, outline=BLACK, width=1)

    for i, header in enumerate(headers[:6]):
        x = col_positions[i] if i < len(col_positions) else x_offset
        w = col_widths[i] if i < len(col_widths) else col_widths[-1]

        bbox = draw.textbbox((0, 0), str(header), font=font_bold)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        text_x = x + (w - text_width) // 2
        text_y = y_offset + (row_heights[2] - text_height) // 2
        draw.text((text_x, text_y), str(header), fill=BLACK, font=font_bold)

        if i < len(col_positions):
            draw.line([(x + w, y_offset), (x + w, y_offset + row_heights[2])], fill=BLACK, width=1)

    y_offset += row_heights[2]

    # 绘制数据行
    for row_idx, row_data in enumerate(data['data_rows']):
        row_height = row_heights[3 + row_idx]
        draw.rectangle([x_offset, y_offset, x_offset + sum(col_widths), y_offset + row_height],
                       fill=WHITE, outline=BLACK, width=1)

        for col_idx, cell_value in enumerate(row_data[:6]):
            x = col_positions[col_idx] if col_idx < len(col_positions) else x_offset
            w = col_widths[col_idx] if col_idx < len(col_widths) else col_widths[-1]

            # 确定字体和颜色（运费和处理费用红色）
            if col_idx in [2, 3]:  # 运费、处理费列
                cell_font = font_red
                cell_color = RED
            else:
                cell_font = font_normal
                cell_color = BLACK

            display_value = str(cell_value) if cell_value else ''

            # 垂直合并处理：非第一行的国家列留空
            if col_idx == 0 and row_idx > 0 and display_value:
                display_value = ''

            bbox = draw.textbbox((0, 0), display_value, font=cell_font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            text_x = x + (w - text_width) // 2
            text_y = y_offset + (row_height - text_height) // 2
            draw.text((text_x, text_y), display_value, fill=cell_color, font=cell_font)

            if col_idx < len(col_positions):
                draw.line([(x + w, y_offset), (x + w, y_offset + row_height)], fill=BLACK, width=1)

        y_offset += row_height

    # 绘制尺寸限制（如果有）
    if len(data['data_rows']) > 0:
        row_height = row_heights[-1]
        draw.rectangle([x_offset, y_offset, x_offset + sum(col_widths), y_offset + row_height],
                       fill=WHITE, outline=BLACK, width=1)

        size_limit = data['data_rows'][0][5] if len(data['data_rows']) > 0 and len(data['data_rows'][0]) > 5 else ''
        if size_limit:
            size_text = '尺寸限制：' + str(size_limit)
            draw.text((x_offset + 10, y_offset + 10), size_text, fill=BLACK, font=font_small)

    # 保存图片
    img.save(output_path, 'PNG', quality=95)
    print(f"图片已生成: {output_path}")
    return output_path


def generate_pricing_images(excel_path, code, output_dir='output'):
    """
    生成指定代码的报价表图片（原内容）

    参数:
        excel_path: Excel文件路径
        code: 产品代码(如'CTD', 'GFD')
        output_dir: 输出目录
    """
    os.makedirs(output_dir, exist_ok=True)

    parser = PricingTableParser(excel_path)

    table = parser.get_table(code)
    if not table:
        print(f"未找到代码 '{code}' 对应的报价表")
        print(f"\n已识别的代码列表:")
        for c in parser.get_all_codes()[:20]:
            print(f"  - {c}")
        return None

    # 生成原始图片
    original_path = os.path.join(output_dir, f'{code}_original.png')
    create_table_image(table, original_path)

    return {
        'original': original_path,
        'data': table
    }


def generate_modified_images(excel_path, code, modifications, output_dir='output'):
    """
    生成修改后的报价表图片

    参数:
        excel_path: Excel文件路径
        code: 产品代码(如'CTD', 'GFD')
        modifications: dict, 包含要修改的值
            {
                'weight_1': '0<W≤0.5',      # 第一重量段
                'freight_1': 85,             # 第一重量段运费
                'process_fee_1': 6,          # 第一重量段处理费
                'weight_2': '0.5<W≤6',      # 第二重量段
                'freight_2': 95,             # 第二重量段运费
                'process_fee_2': 6           # 第二重量段处理费
            }
        output_dir: 输出目录
    """
    import copy

    os.makedirs(output_dir, exist_ok=True)

    parser = PricingTableParser(excel_path)

    table = parser.get_table(code)
    if not table:
        print(f"未找到代码 '{code}' 对应的报价表")
        return None

    # 复制数据
    modified_table = copy.deepcopy(table)

    # 修改数据行
    if modifications and 'data_rows' in modified_table:
        for i, row in enumerate(modified_table['data_rows']):
            if i == 0:  # 第一重量段
                if 'weight_1' in modifications and len(row) > 1:
                    row[1] = modifications['weight_1']
                if 'freight_1' in modifications and len(row) > 2:
                    row[2] = modifications['freight_1']
                if 'process_fee_1' in modifications and len(row) > 3:
                    row[3] = modifications['process_fee_1']
            elif i == 1:  # 第二重量段
                if 'weight_2' in modifications and len(row) > 1:
                    row[1] = modifications['weight_2']
                if 'freight_2' in modifications and len(row) > 2:
                    row[2] = modifications['freight_2']
                if 'process_fee_2' in modifications and len(row) > 3:
                    row[3] = modifications['process_fee_2']

    # 生成修改后的图片
    modified_path = os.path.join(output_dir, f'{code}_modified.png')
    create_table_image(modified_table, modified_path)

    # 同时生成原始图片
    original_path = os.path.join(output_dir, f'{code}_original.png')
    create_table_image(table, original_path)

    return {
        'original': original_path,
        'modified': modified_path,
        'data': modified_table
    }


# ============ 交互式使用示例 ============
if __name__ == '__main__':
    EXCEL_PATH = '出口易物流推广报价表2026.04.22.xlsx'

    print("=" * 70)
    print("出口易物流报价表智能识别工具 v2.0")
    print("=" * 70)
    print("\n功能说明:")
    print("  1. 自动扫描所有工作表，根据表头特征识别数据区域")
    print("  2. 通过产品代码(如CTD、GFD)筛选对应的报价表")
    print("  3. 支持修改重量段、运费、处理费后重新生成对比图片")
    print("  4. 原内容和修改后的内容分别输出为图片")
    print("\n" + "-" * 70)

    # 解析Excel
    parser = PricingTableParser(EXCEL_PATH)

    # 显示已识别的所有报价表
    tables = parser.list_all_tables()
    print(f"\n已识别的报价表 ({len(tables)} 个):")
    print("-" * 70)

    for i, table in enumerate(tables):
        codes = ', '.join(table['codes'][:5])
        name = table['name'][:30] if table['name'] else '未命名'
        print(f"{i+1}. {codes}")
        print(f"   产品: {name}")
        print(f"   工作表: {table['sheet']}")
        print()

    print("-" * 70)

    # 示例：生成指定代码的原始图片
    print("\n【示例1】生成CTD的原始报价表图片:")
    result = generate_pricing_images(EXCEL_PATH, 'CTD', 'output')
    if result:
        print(f"  原始图片: {result['original']}")

    print("\n【示例2】生成GFD的原始报价表图片:")
    result = generate_pricing_images(EXCEL_PATH, 'GFD', 'output')
    if result:
        print(f"  原始图片: {result['original']}")

    print("\n【示例3】修改CTD的运费和处理费后生成对比图片:")
    modifications = {
        'freight_1': 85,      # 第一重量段运费从78改为85
        'freight_2': 95,      # 第二重量段运费从88改为95
        'process_fee_1': 6,   # 处理费从5改为6
    }
    result = generate_modified_images(EXCEL_PATH, 'CTD', modifications, 'output')
    if result:
        print(f"  原始图片: {result['original']}")
        print(f"  修改后图片: {result['modified']}")

    print("\n" + "=" * 70)
    print("运行完成！请查看 output 目录下的图片文件。")
    print("=" * 70)

    # 显示可用代码
    print("\n可用产品代码:")
    all_codes = parser.get_all_codes()
    print(', '.join(all_codes))
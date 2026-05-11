#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按产品代码查询两类报价数据（推广报价 + VIP等级报价），并输出结构化数据和截图。

示例:
python query_product_code.py --code CTD
"""

from __future__ import annotations

import argparse
from copy import deepcopy
from datetime import datetime
import json
import math
import re
import shutil
import traceback
import unicodedata
from pathlib import Path
from time import perf_counter
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from myproject.services.extract_pricing_details import PricingExtractor, ProductLineTable
from myproject.services.grade_quote_extractor import GradeQuoteExtractor, ProductCodeBlock


def _log(msg: str) -> None:
    print(msg, flush=True)


def _ms(start: float) -> int:
    return int((perf_counter() - start) * 1000)


def _find_first_xlsx(workspace: Path, keyword: str) -> Optional[Path]:
    for p in workspace.glob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        if keyword in p.name:
            return p
    return None


def _is_excel_candidate(p: Path) -> bool:
    return p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")


def _find_latest_xlsx_in_dir(folder: Path, keyword: str) -> Optional[Path]:
    candidates = [
        p for p in folder.glob("*.xlsx") if _is_excel_candidate(p) and keyword in p.name
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _iter_batch_dirs_desc(root: Path) -> List[Path]:
    if not root.exists():
        return []
    dirs = [p for p in root.iterdir() if p.is_dir()]
    dirs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return dirs


def _find_latest_xlsx_in_batches(root: Path, keyword: str) -> Optional[Path]:
    candidates: List[Path] = []
    for batch_dir in _iter_batch_dirs_desc(root):
        found = _find_latest_xlsx_in_dir(batch_dir, keyword)
        if found is not None:
            candidates.append(found)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _resolve_latest_pair_from_attachments(
    attachment_root: Path,
    latest_strategy: str,
) -> Tuple[Path, Path, Optional[Path]]:
    promo_keyword = "推广报价表"
    vip_keyword = "直发产品定价"

    if not attachment_root.exists():
        raise FileNotFoundError(f"抓取目录不存在: {attachment_root}")

    if latest_strategy == "batch":
        for batch_dir in _iter_batch_dirs_desc(attachment_root):
            promo = _find_latest_xlsx_in_dir(batch_dir, promo_keyword)
            vip = _find_latest_xlsx_in_dir(batch_dir, vip_keyword)
            if promo is not None and vip is not None:
                return promo, vip, batch_dir
        raise FileNotFoundError(
            "抓取目录中未找到同一批次同时包含推广和VIP报价的Excel文件"
        )

    promo = _find_latest_xlsx_in_batches(attachment_root, promo_keyword)
    vip = _find_latest_xlsx_in_batches(attachment_root, vip_keyword)
    if promo is None:
        raise FileNotFoundError("抓取目录中未找到推广报价Excel")
    if vip is None:
        raise FileNotFoundError("抓取目录中未找到VIP等级报价Excel")
    return promo, vip, None


def _to_abs_path(p: Path, base_dir: Path) -> Path:
    if p.is_absolute():
        return p
    return (base_dir / p).resolve()


def _safe_tag(text: str) -> str:
    s = re.sub(r"[^0-9A-Za-z_.-]+", "_", text).strip("_")
    return s or "run"


def _find_existing_vip_snapshots(snapshot_dir: Path, code: str) -> List[Dict[str, Any]]:
    paths = _find_existing_named_snapshots(snapshot_dir, f"{code.strip().upper()}_VIP")
    out: List[Dict[str, Any]] = []
    for p in paths:
        out.append(
            {
                "产品代码": code.strip().upper(),
                "snapshot": str(p),
                "engine": "cached",
                "reused": True,
            }
        )
    return out


def _find_existing_named_snapshots(snapshot_dir: Path, base_name: str) -> List[Path]:
    if not snapshot_dir.exists():
        return []

    single = snapshot_dir / f"{base_name}.png"
    numbered: List[Tuple[int, Path]] = []
    for p in snapshot_dir.glob(f"{base_name}_*.png"):
        m = re.fullmatch(rf"{re.escape(base_name)}_(\d+).png", p.name)
        if m is None:
            continue
        numbered.append((int(m.group(1)), p))

    if numbered:
        numbered.sort(key=lambda x: x[0])
        return [p for _, p in numbered]
    if single.exists():
        return [single]
    return []


def _normalize_named_snapshots(
    generated_paths: List[str],
    output_dir: Path,
    base_name: str,
) -> List[str]:
    if not generated_paths:
        return []

    output_dir.mkdir(parents=True, exist_ok=True)
    source_paths = [Path(p) for p in generated_paths]
    target_paths: List[Path] = []

    if len(source_paths) == 1:
        target_paths = [output_dir / f"{base_name}.png"]
    else:
        for idx in range(1, len(source_paths) + 1):
            target_paths.append(output_dir / f"{base_name}_{idx:02d}.png")

    for src, dst in zip(source_paths, target_paths):
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)

    # 清理生成过程中的临时命名文件
    target_resolved = {p.resolve() for p in target_paths}
    for src in source_paths:
        try:
            if src.exists() and src.resolve() not in target_resolved:
                src.unlink()
        except Exception:
            continue

    # 清理同前缀的历史文件，避免残留导致读取歧义
    keep_names = {p.name for p in target_paths}
    for p in output_dir.glob(f"{base_name}*.png"):
        is_single = p.name == f"{base_name}.png"
        is_numbered = re.fullmatch(rf"{re.escape(base_name)}_\d+.png", p.name) is not None
        if not (is_single or is_numbered):
            continue
        if p.name not in keep_names:
            try:
                p.unlink()
            except Exception:
                continue

    return [str(p) for p in target_paths]


def _is_prebuilt_payload_compatible(payload: Dict[str, Any], out_dir: Path, code: str, grade: str) -> bool:
    code_u = code.strip().upper()
    grade_u = grade.strip().upper()

    promo = payload.get("推广报价", {}) if isinstance(payload, dict) else {}
    vip = payload.get("VIP等级报价", {}) if isinstance(payload, dict) else {}
    modified = payload.get("推广报价_按等级覆盖后", {}) if isinstance(payload, dict) else {}

    if int(promo.get("matched_count", 0)) > 0:
        if not _find_existing_named_snapshots(out_dir / "promo_snapshots", f"{code_u}_PROMO"):
            return False

    if int(vip.get("matched_block_count", 0)) > 0:
        if not _find_existing_named_snapshots(out_dir / "vip_snapshots", f"{code_u}_VIP"):
            return False

    if not bool(modified.get("skipped", False)) and int(modified.get("matched_table_count", 0)) > 0:
        if not _find_existing_named_snapshots(out_dir / "promo_modified_snapshots", f"{code_u}_{grade_u}_PROMO"):
            return False

    return True


def _infer_source_tag_from_file(file_path: Path, attachment_root: Path) -> Optional[str]:
    try:
        rel = file_path.resolve().relative_to(attachment_root.resolve())
    except Exception:
        return None
    parts = rel.parts
    if len(parts) < 2:
        return None
    return _safe_tag(parts[0])


def _result_json_filename(code: str, grade: str) -> str:
    return f"{code}_{grade}_query_result.json"


def _vip_has_grade_data(records: List[Dict[str, Any]], grade: str) -> bool:
    key = f"{grade}等级"
    for rec in records:
        data = rec.get(key)
        if not isinstance(data, dict):
            continue
        fee_f = data.get("运费")
        fee_h = data.get("处理费")
        if not _is_empty_value(fee_f) or not _is_empty_value(fee_h):
            return True
    return False


def _normalize_code(code: str) -> str:
    return code.strip().upper()


def _normalize_text(text: Any) -> str:
    if text is None:
        return ""
    s = str(text).strip().upper()
    s = s.replace("（", "(").replace("）", ")")
    s = "".join(s.split())
    return s


# 国家匹配可配置规则：支持“推广国家名称 -> VIP分区代码/包含字符串”手动维护。
DEFAULT_COUNTRY_MATCH_RULES: List[Dict[str, List[str]]] = [
    {
        "promo_contains": ["美国"],
        "vip_tokens": ["US", "US1", "US2"],
    },
    {
        "promo_contains": ["秘鲁-利马区域", "秘鲁利马区域", "利马区域"],
        "vip_tokens": ["PE1"],
    },
    {
        "promo_contains": ["秘鲁-非利马区域", "秘鲁非利马区域", "非利马区域"],
        "vip_tokens": ["PE2"],
    },
]

COUNTRY_MATCH_RULES: List[Dict[str, List[str]]] = list(DEFAULT_COUNTRY_MATCH_RULES)


def _load_country_match_rules(config_file: Optional[Path]) -> None:
    global COUNTRY_MATCH_RULES

    if config_file is None:
        COUNTRY_MATCH_RULES = list(DEFAULT_COUNTRY_MATCH_RULES)
        return

    if not config_file.exists():
        COUNTRY_MATCH_RULES = list(DEFAULT_COUNTRY_MATCH_RULES)
        _log(f"[配置] 未找到国家匹配配置文件，使用内置规则: {config_file}")
        return

    raw = json.loads(config_file.read_text(encoding="utf-8"))
    rules = raw.get("rules") if isinstance(raw, dict) else raw
    if not isinstance(rules, list):
        raise ValueError("国家匹配配置文件格式错误：应为 rules 列表或列表根节点")

    normalized_rules: List[Dict[str, List[str]]] = []
    for idx, item in enumerate(rules, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"国家匹配配置第 {idx} 条不是对象")
        promo_contains = item.get("promo_contains", [])
        vip_tokens = item.get("vip_tokens", [])
        if not isinstance(promo_contains, list) or not isinstance(vip_tokens, list):
            raise ValueError(f"国家匹配配置第 {idx} 条格式错误：promo_contains/vip_tokens 必须为数组")

        pc = [str(x).strip() for x in promo_contains if str(x).strip()]
        vt = [str(x).strip() for x in vip_tokens if str(x).strip()]
        if not pc or not vt:
            continue

        normalized_rules.append({"promo_contains": pc, "vip_tokens": vt})

    COUNTRY_MATCH_RULES = normalized_rules or list(DEFAULT_COUNTRY_MATCH_RULES)
    _log(f"[配置] 已加载国家匹配规则: {len(COUNTRY_MATCH_RULES)} 条（{config_file}）")


def _is_empty_value(v: Any) -> bool:
    return v is None or str(v).strip() == ""


def _is_follow_public(v: Any) -> bool:
    return "随公开价" in str(v)


def _build_merged_anchor_lookup(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = (min_row, min_col)
    return lookup


def _set_cell_value_safely(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
) -> bool:
    anchor = merged_lookup.get((row, col))
    if anchor is not None:
        ws.cell(row=anchor[0], column=anchor[1]).value = value
        return (anchor[0], anchor[1]) != (row, col)

    ws.cell(row=row, column=col).value = value
    return False


def _estimate_display_len(text: str) -> int:
    total = 0
    for ch in text:
        total += 2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
    return max(1, total)


def _estimate_row_height(
    ws: Worksheet,
    row: int,
    left_col: int,
    right_col: int,
    merged_lookup: Dict[Tuple[int, int], Tuple[int, int]],
    merged_anchor_col_span: Dict[Tuple[int, int], int],
) -> float:
    max_lines = 1

    for col in range(left_col, right_col + 1):
        anchor = merged_lookup.get((row, col))
        target_row, target_col = anchor if anchor is not None else (row, col)
        # 合并单元格只在锚点行/列计算一次，避免重复放大。
        if (target_row, target_col) != (row, col):
            continue

        value = ws.cell(row=target_row, column=target_col).value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue

        col_span = max(1, merged_anchor_col_span.get((target_row, target_col), 1))
        text_width = 0.0
        for c in range(target_col, min(right_col, target_col + col_span - 1) + 1):
            width = ws.column_dimensions[get_column_letter(c)].width
            text_width += float(width if width is not None else 12)

        # 经验系数：Excel列宽约可容纳 ~1.2 倍字符数（中英文混排做宽字符加权）。
        chars_per_line = max(1, int(text_width * 1.2))
        wrapped_lines = 0
        for line in text.splitlines() or [text]:
            wrapped_lines += max(1, math.ceil(_estimate_display_len(line) / chars_per_line))

        max_lines = max(max_lines, wrapped_lines)

    # 近似Excel默认行高与行间距；限制上限防止异常文本把行高拉太大。
    return min(220.0, 15.0 + max(0, max_lines - 1) * 15.0)


def _autofit_table_rows(ws: Worksheet, table: ProductLineTable, merged_lookup: Dict[Tuple[int, int], Tuple[int, int]]) -> int:
    merged_anchor_col_span: Dict[Tuple[int, int], int] = {}
    for merged_range in ws.merged_cells.ranges:
        merged_anchor_col_span[(merged_range.min_row, merged_range.min_col)] = max(
            merged_anchor_col_span.get((merged_range.min_row, merged_range.min_col), 1),
            merged_range.max_col - merged_range.min_col + 1,
        )

    changed = 0
    for row in range(table.top_row, table.bottom_row + 1):
        required_height = _estimate_row_height(
            ws=ws,
            row=row,
            left_col=table.left_col,
            right_col=table.right_col,
            merged_lookup=merged_lookup,
            merged_anchor_col_span=merged_anchor_col_span,
        )
        current_height = ws.row_dimensions[row].height
        if current_height is None or required_height > float(current_height):
            ws.row_dimensions[row].height = required_height
            changed += 1
    return changed


def _split_country_zone(zone: str) -> List[str]:
    if not zone:
        return []
    tmp = zone.replace("\n", ",")
    raw_tokens = [t.strip() for t in re.split(r"[,，/;；、|]+", tmp) if t.strip()]
    tokens: List[str] = []
    for t in raw_tokens:
        # 去掉括号备注，避免“US2（财务务必两个都设置）”干扰匹配。
        cleaned = re.sub(r"[（(].*?[)）]", "", t).strip()
        if not cleaned:
            cleaned = t.strip()
        if cleaned:
            tokens.append(cleaned.upper())
    return tokens


def _get_config_vip_tokens(promo_country: str) -> List[str]:
    promo_norm = _normalize_text(promo_country)
    if not promo_norm:
        return []

    out: set[str] = set()
    for rule in COUNTRY_MATCH_RULES:
        contains_keys = [_normalize_text(x) for x in rule.get("promo_contains", [])]
        if any(k and k in promo_norm for k in contains_keys):
            for t in rule.get("vip_tokens", []):
                tn = _normalize_text(t)
                if tn:
                    out.add(tn)
    return sorted(out)


def _zone_token_match_alias(token_norm: str, alias_norm: str) -> bool:
    if not token_norm or not alias_norm:
        return False

    if token_norm == alias_norm:
        return True

    # 短英文代码做前缀数字匹配，如 US -> US1/US2。
    if re.fullmatch(r"[A-Z]{2,3}", alias_norm):
        return re.fullmatch(rf"{re.escape(alias_norm)}\d*", token_norm) is not None

    # 长代码允许双向包含，兼容“秘鲁-利马区域”等文本型匹配。
    return alias_norm in token_norm or token_norm in alias_norm


def _country_aliases(country: str) -> List[str]:
    c = _normalize_text(country)
    aliases = {c}
    mapping = {
        "美国": ["US", "USA", "美国"],
        "德国": ["DE", "GERMANY", "德国"],
        "法国": ["FR", "FRANCE", "法国"],
        "西班牙": ["ES", "SPAIN", "西班牙"],
        "葡萄牙": ["PT", "PORTUGAL", "葡萄牙"],
        "荷兰": ["NL", "NETHERLANDS", "荷兰"],
        "英国": ["GB", "UK", "UNITEDKINGDOM", "英国"],
    }
    for k, vals in mapping.items():
        if k in country:
            aliases.update(vals)

    # 可配置映射：推广国家名 -> VIP分区代码。
    aliases.update(_get_config_vip_tokens(country))
    return list(aliases)


def _country_match(promo_country: str, zone_text: str) -> bool:
    zone_norm = _normalize_text(zone_text)
    if "所有国家" in zone_norm:
        return True

    aliases = set(_country_aliases(promo_country))
    tokens = _split_country_zone(zone_text)
    if not tokens:
        return False

    for token in tokens:
        token_norm = _normalize_text(token)
        if any(_zone_token_match_alias(token_norm, _normalize_text(a)) for a in aliases):
            return True
    return False


def _select_grade_fee(vip_record: Dict[str, Any], grade: str) -> Tuple[Any, Any]:
    key = f"{grade}等级"
    data = vip_record.get(key)
    if not isinstance(data, dict):
        return "", ""
    return data.get("运费"), data.get("处理费")


def _strip_parenthesized(text: str) -> str:
    return re.sub(r"[（(].*?[)）]", "", text)


def _parse_weight_range(text: Any) -> Optional[Tuple[float, float]]:
    if text is None:
        return None

    s = str(text).strip()
    if not s:
        return None

    s = _strip_parenthesized(s)
    s = s.replace("KG", "").replace("kg", "")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if len(nums) < 2:
        return None

    low = float(nums[0])
    high = float(nums[1])
    if low > high:
        low, high = high, low
    return low, high


def _weight_match(promo_weight: str, vip_weight: str, tol: float = 0.02) -> bool:
    p_norm = _normalize_text(promo_weight)
    v_norm = _normalize_text(vip_weight)
    if p_norm and p_norm == v_norm:
        return True

    p_range = _parse_weight_range(promo_weight)
    v_range = _parse_weight_range(vip_weight)
    if p_range is None or v_range is None:
        return False

    p_low, p_high = p_range
    v_low, v_high = v_range
    # 1) 边界近似相等
    if abs(p_low - v_low) <= tol and abs(p_high - v_high) <= tol:
        return True

    # 2) 支持区间包含：VIP给大区间(如 0-5)时，可匹配推广子区间(如 0-0.3 / 0.3-0.5)
    promo_in_vip = p_low >= v_low - tol and p_high <= v_high + tol
    vip_in_promo = v_low >= p_low - tol and v_high <= p_high + tol
    return promo_in_vip or vip_in_promo


def _weight_distance(promo_weight: str, vip_weight: str) -> float:
    p_range = _parse_weight_range(promo_weight)
    v_range = _parse_weight_range(vip_weight)
    if p_range is None or v_range is None:
        return 9999.0
    p_low, p_high = p_range
    v_low, v_high = v_range
    return abs(p_low - v_low) + abs(p_high - v_high)


def _pick_best_vip_record(promo_country: str, promo_weight: str, vip_records: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not str(promo_weight).strip():
        return None

    exact: List[Tuple[float, Dict[str, Any]]] = []
    all_country: List[Tuple[float, Dict[str, Any]]] = []

    for rec in vip_records:
        rec_weight_text = str(rec.get("重量段KG", ""))
        if not _weight_match(promo_weight, rec_weight_text):
            continue

        score = _weight_distance(promo_weight, rec_weight_text)

        zone = str(rec.get("国家分区", ""))
        if "所有国家" in _normalize_text(zone):
            all_country.append((score, rec))
            continue

        if _country_match(promo_country, zone):
            exact.append((score, rec))

    if exact:
        exact.sort(key=lambda x: x[0])
        return exact[0][1]
    if all_country:
        all_country.sort(key=lambda x: x[0])
        return all_country[0][1]
    return None


def _promo_table_to_dict(table: ProductLineTable) -> Dict[str, Any]:
    return {
        "sheet": table.sheet_name,
        "产品线": table.product_line,
        "产品代码": table.product_codes,
        "更新时间": table.update_time,
        "header_row": table.header_row,
        "data_start_row": table.data_start_row,
        "data_end_row": table.data_end_row,
        "headers": table.headers,
        "range": {
            "top_left": f"{table.left_col},{table.top_row}",
            "bottom_right": f"{table.right_col},{table.bottom_row}",
        },
        "note": {
            "has_note": table.note_row > 0,
            "row": table.note_row,
            "text": table.note_text,
        },
        "rows": table.rows,
    }


def _default_vip_structured_json(vip_file: Path) -> Path:
    return Path("output") / "cache" / f"{vip_file.stem}_vip_structured.json"


def _default_promo_index_file(promo_file: Path, source_tag: str) -> Path:
    return Path("output") / "index" / source_tag / f"{promo_file.stem}_code_sheet_index.json"


def _default_vip_structured_json_by_batch(vip_file: Path, source_tag: str) -> Path:
    return Path("output") / "cache" / source_tag / f"{vip_file.stem}_vip_structured.json"


def query_promo(
    code: str,
    promo_file: Path,
    snapshot_engine: str,
    out_dir: Path,
    index_file: Optional[Path] = None,
    rebuild_index: bool = False,
    reuse_snapshots: bool = True,
) -> Tuple[Dict[str, Any], List[ProductLineTable]]:
    promo_total_start = perf_counter()
    init_start = perf_counter()
    _log(f"[推广] 开始读取: {promo_file}")
    _log("[推广] 初始化提取器...")
    extractor = PricingExtractor(promo_file)
    init_ms = _ms(init_start)
    _log(f"[推广] 提取器初始化耗时: {init_ms}ms")

    index_start = perf_counter()
    _log("[推广] 提取器初始化完成，开始目录匹配...")
    promo_index_file = index_file or extractor.default_index_file()
    matched_sheets = extractor.get_sheets_by_code(code=code, index_file=promo_index_file, rebuild=rebuild_index)
    anchor_rows_by_sheet = extractor.get_sheet_anchor_rows_by_code(
        code=code,
        index_file=promo_index_file,
        rebuild=rebuild_index,
    )
    if matched_sheets and not anchor_rows_by_sheet and not rebuild_index:
        _log("[推广] 当前索引缺少代码锚点行，自动升级重建索引...")
        anchor_rows_by_sheet = extractor.get_sheet_anchor_rows_by_code(
            code=code,
            index_file=promo_index_file,
            rebuild=True,
        )
    index_ms = _ms(index_start)
    _log(f"[推广] 目录匹配耗时: {index_ms}ms")

    if not matched_sheets:
        total_ms = _ms(promo_total_start)
        return {
            "input_excel": str(promo_file),
            "index_file": str(promo_index_file),
            "matched_sheets": [],
            "matched_count": 0,
            "engine": snapshot_engine,
            "snapshots": [],
            "tables": [],
            "message": f"产品代码 {code} 在推广报价中匹配不到工作表",
            "timing_ms": {
                "init_extractor": init_ms,
                "match_index": index_ms,
                "extract_tables": 0,
                "filter_tables": 0,
                "snapshot": 0,
                "total": total_ms,
            },
        }, []

    _log(f"[推广] 目录命中工作表: {', '.join(matched_sheets)}")
    extract_start = perf_counter()
    tables = extractor.extract(include_sheets=matched_sheets, anchor_rows_by_sheet=anchor_rows_by_sheet)
    extract_ms = _ms(extract_start)
    _log(f"[推广] 工作表提取耗时: {extract_ms}ms")

    filter_start = perf_counter()
    matched = []
    for table in tables:
        codes = [c.upper() for c in table.product_codes]
        if code in codes:
            matched.append(table)
    filter_ms = _ms(filter_start)
    _log(f"[推广] 代码过滤耗时: {filter_ms}ms")

    debug_codes = sorted({c.upper() for table in tables for c in table.product_codes})

    snapshot_dir = out_dir / "promo_snapshots"
    snapshots: List[str] = []
    used_engine = snapshot_engine
    snapshot_ms = 0
    promo_snapshot_base = f"{code}_PROMO"

    if matched:
        existing = _find_existing_named_snapshots(snapshot_dir, promo_snapshot_base) if reuse_snapshots else []
        if existing:
            snapshots = [str(p) for p in existing]
            used_engine = "cached"
            _log(f"[推广] 命中产品线数量: {len(matched)}，复用已有截图: {len(existing)} 张")
        else:
            _log(f"[推广] 命中产品线数量: {len(matched)}，开始生成截图（{snapshot_engine}）")
            snapshot_start = perf_counter()
            if snapshot_engine == "com":
                try:
                    paths = extractor.save_snapshots_com(matched, snapshot_dir)
                except Exception as exc:
                    _log(f"[推广] COM截图失败，自动回退 draw。异常: {type(exc).__name__}: {exc}")
                    _log(f"[推广] COM截图堆栈:\n{traceback.format_exc()}")
                    used_engine = "draw"
                    paths = extractor.save_snapshots(matched, snapshot_dir)
            else:
                paths = extractor.save_snapshots(matched, snapshot_dir)

            snapshots = _normalize_named_snapshots([str(p) for p in paths], snapshot_dir, promo_snapshot_base)
            snapshot_ms = _ms(snapshot_start)
            _log(f"[推广] 截图耗时: {snapshot_ms}ms")

    total_ms = _ms(promo_total_start)
    _log(f"[推广] 完成，匹配 {len(matched)}，截图 {len(snapshots)}，总耗时 {total_ms}ms")

    message = ""
    if not matched:
        if matched_sheets:
            message = (
                f"产品代码 {code} 已命中工作表({', '.join(matched_sheets)})，"
                "但未在可识别的报价表块中匹配到该代码。"
            )
        else:
            message = f"产品代码 {code} 在推广报价中匹配不到工作表"

    return {
        "input_excel": str(promo_file),
        "index_file": str(promo_index_file),
        "matched_sheets": matched_sheets,
        "extracted_table_count": len(tables),
        "extracted_codes": debug_codes,
        "matched_count": len(matched),
        "engine": used_engine,
        "snapshots": snapshots,
        "tables": [_promo_table_to_dict(table) for table in matched],
        "message": message,
        "timing_ms": {
            "init_extractor": init_ms,
            "match_index": index_ms,
            "extract_tables": extract_ms,
            "filter_tables": filter_ms,
            "snapshot": snapshot_ms,
            "total": total_ms,
        },
    }, matched


def query_vip(
    code: str,
    grade: str,
    vip_file: Path,
    snapshot_engine: str,
    out_dir: Path,
    structured_json: Optional[Path] = None,
    rebuild_structured_json: bool = False,
    reuse_snapshots: bool = True,
    skip_snapshot_generation: bool = False,
) -> Dict[str, Any]:
    vip_total_start = perf_counter()
    _log(f"[VIP] 开始读取: {vip_file}")

    structured_file = structured_json or _default_vip_structured_json(vip_file)
    result: Dict[str, Any]
    structured_prepare_ms = 0

    structured_prepare_start = perf_counter()
    if rebuild_structured_json or not structured_file.exists():
        _log(f"[VIP] 开始解析Excel并生成结构化JSON: {structured_file}")
        extractor_for_extract = GradeQuoteExtractor(vip_file, sheet_name="等级报价")
        result = extractor_for_extract.extract()
        structured_file.parent.mkdir(parents=True, exist_ok=True)
        structured_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        _log(f"[VIP] 结构化JSON已生成: {structured_file}")
    else:
        _log(f"[VIP] 使用已有结构化JSON: {structured_file}")
        result = json.loads(structured_file.read_text(encoding="utf-8"))
    structured_prepare_ms = _ms(structured_prepare_start)
    _log(f"[VIP] 结构化数据准备耗时: {structured_prepare_ms}ms")

    blocks_prepare_start = perf_counter()
    extractor = GradeQuoteExtractor(vip_file, sheet_name="等级报价")

    records = [r for r in result["records"] if str(r.get("产品代码", "")).upper() == code]
    by_code = result["by_product_code"].get(code)

    blocks_for_code: List[ProductCodeBlock] = []
    for b in result.get("blocks", []):
        code_list = [str(x).upper() for x in b.get("产品代码列表", [])]
        if code in code_list:
            blocks_for_code.append(
                ProductCodeBlock(
                    product_name=b.get("产品名称", ""),
                    service_text=b.get("服务代码原文", ""),
                    service_codes=[code],
                    start_row=int(b.get("start_row", 0)),
                    end_row=int(b.get("end_row", 0)),
                )
            )

    header_row = int(result["meta"]["header_row"])
    cm = extractor._build_column_map(header_row)
    prepare_blocks_ms = _ms(blocks_prepare_start)
    _log(f"[VIP] 代码筛选与块准备耗时: {prepare_blocks_ms}ms")

    snapshot_dir = out_dir / "vip_snapshots"
    if skip_snapshot_generation:
        existing = _find_existing_vip_snapshots(snapshot_dir, code)
        if existing:
            _log(f"[VIP] 跳过截图生成，复用已有截图: {len(existing)} 张")
            total_ms = _ms(vip_total_start)
            return {
                "input_excel": str(vip_file),
                "structured_json": str(structured_file),
                "grade": grade,
                "matched_record_count": len(records),
                "matched_block_count": len(blocks_for_code),
                "records": records,
                "by_product_code": by_code,
                "snapshots": existing,
                "timing_ms": {
                    "prepare_structured": structured_prepare_ms,
                    "prepare_blocks": prepare_blocks_ms,
                    "snapshot": 0,
                    "total": total_ms,
                },
            }

    _log(f"[VIP] 命中记录 {len(records)}，命中块 {len(blocks_for_code)}，开始生成截图（{snapshot_engine}）")
    snapshot_start = perf_counter()
    snapshot_items = extractor.save_block_snapshots(
        blocks=blocks_for_code,
        header_row=header_row,
        cm=cm,
        output_dir=snapshot_dir,
        engine=snapshot_engine,
        reuse_existing=reuse_snapshots,
    )
    snapshot_ms = _ms(snapshot_start)
    _log(f"[VIP] 截图耗时: {snapshot_ms}ms")

    total_ms = _ms(vip_total_start)
    _log(f"[VIP] 完成，截图 {len(snapshot_items)}，总耗时 {total_ms}ms")

    return {
        "input_excel": str(vip_file),
        "structured_json": str(structured_file),
        "grade": grade,
        "matched_record_count": len(records),
        "matched_block_count": len(blocks_for_code),
        "records": records,
        "by_product_code": by_code,
        "snapshots": snapshot_items,
        "timing_ms": {
            "prepare_structured": structured_prepare_ms,
            "prepare_blocks": prepare_blocks_ms,
            "snapshot": snapshot_ms,
            "total": total_ms,
        },
    }


def apply_grade_to_promo(
    code: str,
    grade: str,
    promo_file: Path,
    vip_records: List[Dict[str, Any]],
    snapshot_engine: str,
    out_dir: Path,
    include_sheets: Optional[List[str]] = None,
    source_tables: Optional[List[ProductLineTable]] = None,
    reuse_snapshots: bool = True,
) -> Dict[str, Any]:
    apply_total_start = perf_counter()
    _log(f"[覆盖] 开始按等级 {grade} 覆盖推广报价")
    if source_tables is not None:
        tables = source_tables
        extract_source_ms = 0
        _log("[覆盖] 复用推广阶段已提取表，跳过原始表提取")
    else:
        extract_source_start = perf_counter()
        source_extractor = PricingExtractor(promo_file)
        tables = [
            table_item
            for table_item in source_extractor.extract(include_sheets=include_sheets)
            if code in [x.upper() for x in table_item.product_codes]
        ]
        extract_source_ms = _ms(extract_source_start)
        _log(f"[覆盖] 原始表提取耗时: {extract_source_ms}ms")

    match_rows_start = perf_counter()
    update_logs: List[Dict[str, Any]] = []
    for table in tables:
        for row in table.rows:
            row_no = int(row.get("行号", 0))
            promo_country = str(row.get("国家", ""))
            promo_weight = str(row.get("重量段/KG", ""))
            is_data_row = _parse_weight_range(promo_weight) is not None

            rec = _pick_best_vip_record(promo_country, promo_weight, vip_records)
            current_f = row.get("运费（RMB/KG）")
            current_h = row.get("处理费(RMB/票)")

            new_f = ""
            new_h = ""
            status = "unmatched_blank"
            matched_zone = ""
            matched_vip_weight = ""

            if not is_data_row:
                status = "skip_non_data_row"
                update_logs.append(
                    {
                        "sheet": table.sheet_name,
                        "产品线": table.product_line,
                        "row": row_no,
                        "国家": promo_country,
                        "重量段KG": promo_weight,
                        "原运费": current_f,
                        "原处理费": current_h,
                        "新运费": current_f,
                        "新处理费": current_h,
                        "匹配状态": status,
                        "匹配国家分区": matched_zone,
                        "匹配VIP重量段KG": matched_vip_weight,
                        "should_write": False,
                    }
                )
                continue

            if rec is not None:
                fee_f, fee_h = _select_grade_fee(rec, grade)
                matched_zone = str(rec.get("国家分区", ""))
                matched_vip_weight = str(rec.get("重量段KG", ""))
                status = "matched"

                if _is_follow_public(fee_f):
                    new_f = current_f
                    status = "follow_public_keep"
                elif _is_empty_value(fee_f):
                    new_f = ""
                else:
                    new_f = fee_f

                if _is_follow_public(fee_h):
                    new_h = current_h
                    status = "follow_public_keep"
                elif _is_empty_value(fee_h):
                    new_h = ""
                else:
                    new_h = fee_h

            update_logs.append(
                {
                    "sheet": table.sheet_name,
                    "产品线": table.product_line,
                    "row": row_no,
                    "国家": promo_country,
                    "重量段KG": promo_weight,
                    "原运费": current_f,
                    "原处理费": current_h,
                    "新运费": new_f,
                    "新处理费": new_h,
                    "匹配状态": status,
                    "匹配国家分区": matched_zone,
                    "匹配VIP重量段KG": matched_vip_weight,
                    "should_write": True,
                }
            )
    match_rows_ms = _ms(match_rows_start)
    _log(f"[覆盖] 行匹配与更新计算耗时: {match_rows_ms}ms")

    # 在复制文件上覆盖值，保留原文件不变
    write_excel_start = perf_counter()
    modified_excel = out_dir / f"{code}_promo_modified_by_{grade}.xlsx"
    shutil.copy2(promo_file, modified_excel)
    wb = load_workbook(modified_excel)
    merged_lookup_by_sheet: Dict[str, Dict[Tuple[int, int], Tuple[int, int]]] = {}
    merged_redirect_count = 0
    touched_table_keys: set[Tuple[str, str]] = set()

    for log in update_logs:
        if not bool(log.get("should_write", True)):
            continue

        ws = wb[log["sheet"]]
        row_no = int(log["row"])
        merged_lookup = merged_lookup_by_sheet.get(ws.title)
        if merged_lookup is None:
            merged_lookup = _build_merged_anchor_lookup(ws)
            merged_lookup_by_sheet[ws.title] = merged_lookup

        target_table = None
        for table_item in tables:
            if table_item.sheet_name == log["sheet"] and table_item.product_line == log["产品线"]:
                target_table = table_item
                break
        if target_table is None:
            continue
        touched_table_keys.add((target_table.sheet_name, target_table.product_line))

        freight_col = target_table.left_col + 2
        handling_col = target_table.left_col + 3

        if _set_cell_value_safely(ws, row_no, freight_col, log["新运费"], merged_lookup):
            merged_redirect_count += 1
        if _set_cell_value_safely(ws, row_no, handling_col, log["新处理费"], merged_lookup):
            merged_redirect_count += 1

    autofit_changed_rows = 0
    table_map: Dict[Tuple[str, str], ProductLineTable] = {
        (t.sheet_name, t.product_line): t for t in tables
    }
    for key in touched_table_keys:
        table_item = table_map.get(key)
        if table_item is None:
            continue
        ws = wb[table_item.sheet_name]
        merged_lookup = merged_lookup_by_sheet.get(ws.title)
        if merged_lookup is None:
            merged_lookup = _build_merged_anchor_lookup(ws)
            merged_lookup_by_sheet[ws.title] = merged_lookup
        autofit_changed_rows += _autofit_table_rows(ws, table_item, merged_lookup)

    # 覆盖输出仅保留本次命中的相关工作表，避免夹带不相干sheet。
    related_sheet_names = {table_item.sheet_name for table_item in tables}
    if related_sheet_names:
        removed_sheet_names: List[str] = []
        for ws in list(wb.worksheets):
            if ws.title not in related_sheet_names:
                wb.remove(ws)
                removed_sheet_names.append(ws.title)
        if removed_sheet_names:
            _log(f"[覆盖] 已移除无关工作表: {len(removed_sheet_names)}")
        _log(f"[覆盖] 保留工作表: {', '.join(sorted(related_sheet_names))}")

    wb.save(modified_excel)
    write_excel_ms = _ms(write_excel_start)
    if autofit_changed_rows > 0:
        _log(f"[覆盖] 行高自适应调整行数: {autofit_changed_rows}")
    if merged_redirect_count > 0:
        _log(f"[覆盖] 合并单元格写入重定向次数: {merged_redirect_count}")
    _log(f"[覆盖] 覆盖写回Excel耗时: {write_excel_ms}ms")
    _log(f"[覆盖] 已写入覆盖结果Excel: {modified_excel}")

    # 复用原始提取结果构建覆盖后结构化数据，避免再次扫描整表。
    log_map: Dict[Tuple[str, str, int], Dict[str, Any]] = {}
    for item in update_logs:
        key = (str(item["sheet"]), str(item["产品线"]), int(item["row"]))
        log_map[key] = item

    modified_tables: List[ProductLineTable] = deepcopy(tables)
    for table_item in modified_tables:
        for row in table_item.rows:
            row_no = int(row.get("行号", 0))
            key = (table_item.sheet_name, table_item.product_line, row_no)
            hit = log_map.get(key)
            if hit is None:
                continue
            row["运费（RMB/KG）"] = hit.get("新运费", "")
            row["处理费(RMB/票)"] = hit.get("新处理费", "")

    reextract_ms = 0
    _log("[覆盖] 复用提取结果构建覆盖后结构，跳过覆盖后二次提取")

    snapshot_dir = out_dir / "promo_modified_snapshots"
    used_engine = snapshot_engine
    modified_snapshot_base = f"{code}_{grade}_PROMO"
    existing_modified = _find_existing_named_snapshots(snapshot_dir, modified_snapshot_base) if reuse_snapshots else []
    if existing_modified:
        snap_paths = existing_modified
        used_engine = "cached"
        snapshot_ms = 0
        _log(f"[覆盖] 复用已有覆盖截图: {len(existing_modified)} 张")
    else:
        _log(f"[覆盖] 开始生成覆盖后截图（{snapshot_engine}）")
        snapshot_start = perf_counter()
        modified_extractor = PricingExtractor(modified_excel)
        if snapshot_engine == "com":
            try:
                raw_paths = modified_extractor.save_snapshots_com(modified_tables, snapshot_dir)
            except Exception as exc:
                _log(f"[覆盖] COM截图失败，自动回退 draw。异常: {type(exc).__name__}: {exc}")
                _log(f"[覆盖] COM截图堆栈:\n{traceback.format_exc()}")
                used_engine = "draw"
                raw_paths = modified_extractor.save_snapshots(modified_tables, snapshot_dir)
        else:
            raw_paths = modified_extractor.save_snapshots(modified_tables, snapshot_dir)
        snap_paths = [Path(p) for p in _normalize_named_snapshots([str(p) for p in raw_paths], snapshot_dir, modified_snapshot_base)]
        snapshot_ms = _ms(snapshot_start)
        _log(f"[覆盖] 截图耗时: {snapshot_ms}ms")

    total_ms = _ms(apply_total_start)
    _log(f"[覆盖] 完成，截图 {len(snap_paths)}，总耗时 {total_ms}ms")

    return {
        "grade": grade,
        "modified_excel": str(modified_excel),
        "matched_table_count": len(modified_tables),
        "update_logs": update_logs,
        "modified_tables": [_promo_table_to_dict(table_item) for table_item in modified_tables],
        "snapshots": [str(p) for p in snap_paths],
        "engine": used_engine,
        "timing_ms": {
            "extract_source": extract_source_ms,
            "match_rows": match_rows_ms,
            "write_excel": write_excel_ms,
            "reextract_modified": reextract_ms,
            "snapshot": snapshot_ms,
            "total": total_ms,
        },
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按产品代码查询推广报价和VIP等级报价")
    parser.add_argument("--code", type=str, required=True, help="产品代码，如 CTD")
    parser.add_argument("--grade", choices=["A", "B", "C", "D", "E", "F"], required=True, help="等级报价档位")
    parser.add_argument(
        "--promo-input",
        type=Path,
        default=None,
        help="推广报价Excel路径（可选；为空时自动从抓取目录选择最新文件）",
    )
    parser.add_argument(
        "--vip-input",
        type=Path,
        default=None,
        help="VIP等级报价Excel路径（可选；为空时自动从抓取目录选择最新文件）",
    )
    parser.add_argument(
        "--attachment-root",
        type=Path,
        default=Path("output") / "mail_attachments" / "direct_price_adjustment",
        help="抓取附件根目录，默认 output/mail_attachments/direct_price_adjustment",
    )
    parser.add_argument(
        "--latest-strategy",
        choices=["batch", "file"],
        default="batch",
        help="最新文件选择策略：batch=同批次成对文件优先；file=分别取全局最新",
    )
    parser.add_argument(
        "--source-tag",
        type=str,
        default=None,
        help="可选：显式指定本次输出与缓存批次标识（用于预生成任务）",
    )
    parser.add_argument(
        "--force-regenerate",
        action="store_true",
        help="强制重建，不使用已预生成的同 code+grade 结果",
    )
    parser.add_argument(
        "--skip-vip-snapshot-generation",
        action="store_true",
        help="跳过VIP截图生成；若已有同产品代码截图则直接复用",
    )
    parser.add_argument(
        "--snapshot-engine",
        choices=["com", "draw"],
        default="com",
        help="截图引擎，默认 com；失败时自动回退 draw",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output") / "code_query",
        help="输出目录",
    )
    parser.add_argument(
        "--promo-index-file",
        type=Path,
        default=None,
        help="推广报价产品代码-工作表目录文件路径；为空时使用默认路径",
    )
    parser.add_argument(
        "--vip-structured-json",
        type=Path,
        default=None,
        help="VIP价格设置结构化JSON路径；存在则直接复用，不存在时自动生成",
    )
    parser.add_argument(
        "--rebuild-vip-structured-json",
        action="store_true",
        help="强制重建VIP价格设置结构化JSON",
    )
    parser.add_argument(
        "--country-match-config",
        type=Path,
        default=Path("country_match_rules.json"),
        help="国家匹配配置文件(JSON)，用于维护推广国家与VIP分区代码映射；不存在时使用内置规则",
    )
    parser.add_argument(
        "--rebuild-promo-index",
        action="store_true",
        help="强制重建推广报价产品代码-工作表目录",
    )
    return parser


def main() -> None:
    main_total_start = perf_counter()
    args = build_parser().parse_args()
    _load_country_match_rules(args.country_match_config)
    code = _normalize_code(args.code)
    grade = _normalize_code(args.grade)

    workspace = Path.cwd()

    promo_file = _to_abs_path(args.promo_input, workspace) if args.promo_input else None
    vip_file = _to_abs_path(args.vip_input, workspace) if args.vip_input else None
    batch_dir: Optional[Path] = None

    if promo_file is None or vip_file is None:
        attachment_root = _to_abs_path(args.attachment_root, workspace)
        auto_promo, auto_vip, auto_batch_dir = _resolve_latest_pair_from_attachments(
            attachment_root=attachment_root,
            latest_strategy=args.latest_strategy,
        )
        if promo_file is None:
            promo_file = auto_promo
        if vip_file is None:
            vip_file = auto_vip
        batch_dir = auto_batch_dir

    if promo_file is None:
        raise FileNotFoundError("未找到推广报价Excel，请通过 --promo-input 指定或检查抓取目录")
    if vip_file is None:
        raise FileNotFoundError("未找到VIP等级报价Excel，请通过 --vip-input 指定或检查抓取目录")
    if not promo_file.exists():
        raise FileNotFoundError(f"推广报价Excel不存在: {promo_file}")
    if not vip_file.exists():
        raise FileNotFoundError(f"VIP等级报价Excel不存在: {vip_file}")

    attachment_root = _to_abs_path(args.attachment_root, workspace)
    inferred_tag = _infer_source_tag_from_file(promo_file, attachment_root) or _infer_source_tag_from_file(vip_file, attachment_root)
    if args.source_tag:
        source_tag = _safe_tag(args.source_tag)
    elif batch_dir is not None:
        source_tag = _safe_tag(batch_dir.name)
    elif inferred_tag:
        source_tag = inferred_tag
    else:
        source_tag = _safe_tag(f"manual_{datetime.now().strftime('%Y%m%d%H%M%S')}")

    out_dir = args.output_dir / source_tag / code
    out_dir.mkdir(parents=True, exist_ok=True)
    output_json = out_dir / _result_json_filename(code, grade)

    if output_json.exists() and not args.force_regenerate:
        payload = json.loads(output_json.read_text(encoding="utf-8"))
        if _is_prebuilt_payload_compatible(payload, out_dir, code, grade):
            _log(f"[命中] 发现已预生成结果，直接返回: {output_json}")
            _log(f"输出JSON: {output_json}")
            return
        _log(f"[命中] 发现旧格式预生成结果，自动重建: {output_json}")

    effective_promo_index_file = (
        _to_abs_path(args.promo_index_file, workspace)
        if args.promo_index_file is not None
        else _to_abs_path(_default_promo_index_file(promo_file, source_tag), workspace)
    )
    effective_promo_index_file.parent.mkdir(parents=True, exist_ok=True)

    effective_vip_structured_json = (
        _to_abs_path(args.vip_structured_json, workspace)
        if args.vip_structured_json is not None
        else _to_abs_path(_default_vip_structured_json_by_batch(vip_file, source_tag), workspace)
    )
    effective_vip_structured_json.parent.mkdir(parents=True, exist_ok=True)

    _log(f"[开始] 产品代码={code}, 等级={grade}, 截图引擎={args.snapshot_engine}")
    _log(f"[输入] 推广文件: {promo_file}")
    _log(f"[输入] VIP文件: {vip_file}")
    _log(f"[输入] 来源标识: {source_tag}")
    _log(f"[输入] 推广索引文件: {effective_promo_index_file}")
    _log(f"[输入] VIP结构化JSON: {effective_vip_structured_json}")

    promo_stage_start = perf_counter()
    promo_result, promo_tables = query_promo(
        code,
        promo_file,
        args.snapshot_engine,
        out_dir,
        index_file=effective_promo_index_file,
        rebuild_index=args.rebuild_promo_index,
        reuse_snapshots=not args.force_regenerate,
    )
    promo_ms = _ms(promo_stage_start)
    _log(f"[耗时] 推广阶段总耗时: {promo_ms}ms")

    if promo_result.get("matched_count", 0) <= 0:
        stop_message = promo_result.get("message") or f"产品代码 {code} 在推广报价中匹配不到工作表，已停止后续执行。"
        payload = {
            "产品代码": code,
            "等级": grade,
            "输入来源": {
                "promo_file": str(promo_file),
                "vip_file": str(vip_file),
                "batch_dir": str(batch_dir) if batch_dir is not None else "",
                "source_tag": source_tag,
            },
            "推广报价": promo_result,
            "message": stop_message,
            "timing_ms": {
                "promo": promo_ms,
                "vip": 0,
                "apply_grade": 0,
                "total": _ms(main_total_start),
            },
        }
        output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _log(payload["message"])
        _log(f"输出JSON: {output_json}")
        _log(f"[耗时] 总耗时: {_ms(main_total_start)}ms")
        return

    vip_stage_start = perf_counter()
    vip_result = query_vip(
        code,
        grade,
        vip_file,
        args.snapshot_engine,
        out_dir,
        structured_json=effective_vip_structured_json,
        rebuild_structured_json=args.rebuild_vip_structured_json,
        reuse_snapshots=not args.force_regenerate,
        skip_snapshot_generation=args.skip_vip_snapshot_generation,
    )
    vip_ms = _ms(vip_stage_start)
    _log(f"[耗时] VIP阶段总耗时: {vip_ms}ms")

    apply_stage_start = perf_counter()
    if not _vip_has_grade_data(vip_result.get("records", []), grade):
        _log(f"[覆盖] structured中未找到产品代码 {code} 的 {grade} 等级有效数据，跳过覆盖与 promo_modified_snapshots 生成")
        modified_result = {
            "grade": grade,
            "modified_excel": "",
            "matched_table_count": 0,
            "update_logs": [],
            "modified_tables": [],
            "snapshots": [],
            "engine": args.snapshot_engine,
            "skipped": True,
            "skip_reason": f"structured.json 中缺少产品代码 {code} 的 {grade} 等级有效数据",
            "timing_ms": {
                "extract_source": 0,
                "match_rows": 0,
                "write_excel": 0,
                "reextract_modified": 0,
                "snapshot": 0,
                "total": 0,
            },
        }
        apply_ms = 0
    else:
        modified_result = apply_grade_to_promo(
            code=code,
            grade=grade,
            promo_file=promo_file,
            vip_records=vip_result["records"],
            snapshot_engine=args.snapshot_engine,
            out_dir=out_dir,
            include_sheets=promo_result.get("matched_sheets"),
            source_tables=promo_tables,
            reuse_snapshots=not args.force_regenerate,
        )
        apply_ms = _ms(apply_stage_start)
    _log(f"[耗时] 覆盖阶段总耗时: {apply_ms}ms")

    total_ms = _ms(main_total_start)

    payload = {
        "产品代码": code,
        "等级": grade,
        "输入来源": {
            "promo_file": str(promo_file),
            "vip_file": str(vip_file),
            "batch_dir": str(batch_dir) if batch_dir is not None else "",
            "source_tag": source_tag,
        },
        "推广报价": promo_result,
        "VIP等级报价": vip_result,
        "推广报价_按等级覆盖后": modified_result,
        "timing_ms": {
            "promo": promo_ms,
            "vip": vip_ms,
            "apply_grade": apply_ms,
            "total": total_ms,
        },
    }

    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"产品代码: {code}")
    print(f"等级: {grade}")
    print(f"推广索引文件: {promo_result.get('index_file', '')}")
    print(f"推广命中工作表: {', '.join(promo_result.get('matched_sheets', []))}")
    print(f"推广报价匹配: {promo_result['matched_count']}")
    print(f"VIP记录匹配: {vip_result['matched_record_count']}")
    print(f"VIP块匹配: {vip_result['matched_block_count']}")
    print(f"覆盖后推广表匹配: {modified_result['matched_table_count']}")
    print(f"耗时(推广/VIP/覆盖/总计, ms): {promo_ms}/{vip_ms}/{apply_ms}/{total_ms}")
    print(f"输出JSON: {output_json}")
    print(f"推广截图目录: {out_dir / 'promo_snapshots'}")
    print(f"VIP截图目录: {out_dir / 'vip_snapshots'}")
    print(f"覆盖后推广截图目录: {out_dir / 'promo_modified_snapshots'}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        _log(f"[错误] {type(exc).__name__}: {exc}")
        _log(traceback.format_exc())
        raise
